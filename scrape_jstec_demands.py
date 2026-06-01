#!/usr/bin/env python3
"""
Scrape technical demand records from 江苏省技术产权交易市场.

Example:
    python scrape_jstec_demands.py --output jstec_demands.xlsx
    python scrape_jstec_demands.py --limit 20 --jsonl jstec_demands.jsonl
    python scrape_jstec_demands.py --login --output jstec_demands.xlsx
"""

from __future__ import annotations

import argparse
import csv
import getpass
import html
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Tuple

import requests


BASE_URL = "https://www.jstec.com.cn"
LIST_API = f"{BASE_URL}/portal/crowdsourceDemand/getDemandPageList"
DETAIL_API = f"{BASE_URL}/portal/crowdsourceDemand/getDemandDetail"
LOGIN_API = f"{BASE_URL}/portal/login"
CAPTCHA_API = f"{BASE_URL}/portal/getCaptchaCode"
AREA_JS_URL = f"{BASE_URL}/_nuxt/e760a04.js"
DETAIL_PAGE_URL = f"{BASE_URL}/crowdsourcing/demandDetail?id={{id}}"


LIST_DEFAULT_QUERY = {
    "status": "",
    "keyword": "",
    "order": "",
    "sort": "",
    "highCooperation": "",
    "cooperationMode": "",
    "priceRange": "",
    "type": "",
    "publishYear": "",
    "company": "",
    "area": "",
}


COOPERATION_MODE_MAP = {
    "00": "股权投资",
    "01": "技术转让",
    "02": "许可使用",
    "03": "合作开发",
    "04": "合作兴办新企业",
    "05": "其他",
    "06": "技术咨询",
    "07": "技术服务",
    "20": "银行贷款",
    "21": "股权融资",
}


DEMAND_TYPE_MAP = {
    "00": "关键技术研发",
    "01": "产品升级",
    "02": "技术改造",
    "03": "设备改进",
    "04": "技术交易",
    "05": "技术咨询",
    "06": "海外创新合作",
    "07": "科技金融",
    "10": "测试参数",
    "11": "测试方法",
    "12": "测试设备",
    "13": "测试标准",
    "99": "其他",
}


CSV_FIELDS = [
    "需求名称",
    "需求编号",
    "合作方式",
    "意向投入",
    "联系方式",
    "发布者",
    "需求详情",
    "技术领域",
    "需求类型",
    "所在地区",
    "需求ID",
    "详情页链接",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    # Excel/openpyxl cannot write most ASCII control characters.
    text = re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p\s*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def label_with_code(code: Any, mapping: Dict[str, str]) -> str:
    value = clean_text(code)
    if not value:
        return ""
    return mapping.get(value, value)


def format_intended_price(value: Any) -> str:
    return clean_text(value)


def is_masked_contact(value: str) -> bool:
    return bool(re.search(r"[xX*＊]", value))


def choose_contact(demand: Dict[str, Any]) -> str:
    tech_manager = demand.get("techManagerOpenDto")
    if not isinstance(tech_manager, dict):
        tech_manager = {}

    candidates = [
        clean_text(demand.get("phone")),
        clean_text(demand.get("publisherPhone")),
        clean_text(tech_manager.get("phone")),
        clean_text(demand.get("userName")),
    ]
    candidates = [value for value in candidates if value]

    for value in candidates:
        if not is_masked_contact(value):
            return value
    return candidates[0] if candidates else ""


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
            "Referer": f"{BASE_URL}/crowdsourcing/demandList",
            "area_code": "jiangsu",
        }
    )
    return session


def request_json(
    session: requests.Session,
    url: str,
    *,
    params: Optional[Dict[str, Any]] = None,
    timeout: float = 20,
    retries: int = 6,
) -> Dict[str, Any]:
    last_error: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, params=params, timeout=timeout)
            response.raise_for_status()
            data = response.json()
            if data.get("code") != "0000":
                raise RuntimeError(f"接口返回异常：{data.get('code')} {data.get('message')}")
            return data
        except Exception as exc:  # noqa: BLE001 - retry any network/API parsing failure.
            last_error = exc
            if attempt < retries:
                sleep_seconds = min(2.0 * attempt, 30.0)
                print(
                    f"请求失败，{sleep_seconds:.0f} 秒后重试 "
                    f"({attempt}/{retries})：{url} params={params!r} error={exc}",
                    file=sys.stderr,
                )
                time.sleep(sleep_seconds)
    raise RuntimeError(f"请求失败：{url} params={params!r} error={last_error}") from last_error


def login_if_requested(session: requests.Session, args: argparse.Namespace) -> None:
    username = args.username
    password = args.password
    verify_code = args.verify_code

    should_login = args.login or username or password or verify_code
    if not should_login:
        return

    if not username:
        username = input("请输入 JSTEC 用户名/手机号：").strip()
    if not password:
        password = getpass.getpass("请输入 JSTEC 密码：")

    if not verify_code:
        captcha_path = Path(args.captcha_path)
        response = session.get(f"{CAPTCHA_API}?{int(time.time() * 1000)}", timeout=args.timeout)
        response.raise_for_status()
        captcha_path.write_bytes(response.content)
        print(f"验证码图片已保存：{captcha_path.resolve()}", file=sys.stderr)
        verify_code = input("请输入验证码图片中的字符：").strip()

    payload = {
        "userName": username,
        "password": password,
        "verifyCode": verify_code,
    }
    response = session.post(LOGIN_API, json=payload, timeout=args.timeout)
    response.raise_for_status()
    data = response.json()
    if data.get("code") != "0000":
        raise RuntimeError(f"登录失败：{data.get('code')} {data.get('message')}")
    print("登录成功，将使用当前会话抓取详情。", file=sys.stderr)


def fetch_area_code_map(session: requests.Session, timeout: float) -> Dict[str, str]:
    """Extract area code labels from the site's own bundled JS."""
    try:
        response = session.get(AREA_JS_URL, timeout=timeout)
        response.raise_for_status()
    except requests.RequestException as exc:
        print(f"地区代码表读取失败，将保留原始编码：{exc}", file=sys.stderr)
        return {}

    area_map: Dict[str, str] = {}
    for raw_code, name in re.findall(r"([0-9]+(?:e[0-9]+)?):\"([^\"]+)\"", response.text):
        try:
            code = str(int(float(raw_code)))
        except ValueError:
            continue
        if len(code) == 6:
            area_map[code] = name
    return area_map


def format_region(region: Any, area_map: Dict[str, str]) -> str:
    value = clean_text(region)
    if not value:
        return ""
    if "," not in value:
        return area_map.get(value, value)
    names = [area_map.get(code.strip(), code.strip()) for code in value.split(",") if code.strip()]
    return " / ".join(names)


def format_detail_text(demand: Dict[str, Any]) -> str:
    text = clean_text(demand.get("summary"))
    overall = demand.get("overallDto")
    if isinstance(overall, dict):
        extra = clean_text(overall.get("description"))
        if extra and extra not in text:
            text = f"{text}\n{extra}" if text else extra
    return text


def extract_row(demand: Dict[str, Any], area_map: Dict[str, str]) -> Dict[str, str]:
    demand_id = clean_text(demand.get("id"))
    overall = demand.get("overallDto") if isinstance(demand.get("overallDto"), dict) else {}
    technology_field = (
        clean_text(demand.get("highCooperationName"))
        or clean_text(demand.get("highCooperation"))
        or clean_text(overall.get("technologyName"))
    )

    return {
        "需求名称": clean_text(demand.get("name")),
        "需求编号": clean_text(demand.get("no")),
        "合作方式": label_with_code(demand.get("cooperationMode"), COOPERATION_MODE_MAP),
        "意向投入": format_intended_price(demand.get("intendedPrice")),
        "联系方式": choose_contact(demand),
        "发布者": clean_text(demand.get("publisher")),
        "需求详情": format_detail_text(demand),
        "技术领域": technology_field,
        "需求类型": label_with_code(demand.get("type"), DEMAND_TYPE_MAP),
        "所在地区": format_region(demand.get("region"), area_map),
        "需求ID": demand_id,
        "详情页链接": DETAIL_PAGE_URL.format(id=demand_id) if demand_id else "",
    }


def fetch_list_page(
    session: requests.Session,
    *,
    page: int,
    page_size: int,
    timeout: float,
) -> Tuple[list[Dict[str, Any]], int]:
    params = dict(LIST_DEFAULT_QUERY)
    params.update({"page": page, "row": page_size})
    payload = request_json(session, LIST_API, params=params, timeout=timeout, retries=2)
    data = payload.get("data") or {}
    return data.get("rows") or [], int(data.get("total") or 0)


def fetch_list_page_one_by_one(
    session: requests.Session,
    *,
    page: int,
    page_size: int,
    timeout: float,
    known_total: Optional[int] = None,
    skip_log_path: Optional[Path] = None,
) -> Tuple[list[Dict[str, Any]], int]:
    """Fallback for server errors on page=N,row=10: fetch the same window one row at a time."""
    rows: list[Dict[str, Any]] = []
    total = int(known_total or 0)
    start_index = (page - 1) * page_size + 1
    end_index = start_index + page_size - 1

    print(
        f"第 {page} 页批量接口失败，改用单条模式抓取第 {start_index}-{end_index} 条。",
        file=sys.stderr,
    )
    for index in range(start_index, end_index + 1):
        if total and index > total:
            break
        try:
            item_rows, item_total = fetch_list_page(
                session,
                page=index,
                page_size=1,
                timeout=timeout,
            )
        except Exception as exc:  # noqa: BLE001 - continue with the next absolute row.
            message = f"第 {index} 条列表记录读取失败，已跳过：{exc}"
            print(message, file=sys.stderr)
            append_skip_log(skip_log_path, message)
            continue

        if item_total:
            total = item_total
        if item_rows:
            rows.extend(item_rows[:1])
        if timeout > 0:
            time.sleep(0.2)

    return rows, total


def fetch_detail(
    session: requests.Session,
    demand_id: str,
    *,
    timeout: float,
) -> Dict[str, Any]:
    payload = request_json(
        session,
        DETAIL_API,
        params={"demandId": demand_id, "activityId": ""},
        timeout=timeout,
    )
    data = payload.get("data")
    if not isinstance(data, dict):
        raise RuntimeError(f"详情接口数据格式异常：{demand_id}")
    return data


def append_skip_log(skip_log_path: Optional[Path], message: str) -> None:
    if not skip_log_path:
        return
    skip_log_path.parent.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    with skip_log_path.open("a", encoding="utf-8") as file:
        file.write(f"[{timestamp}] {message}\n")


def iter_demands(
    session: requests.Session,
    *,
    area_map: Dict[str, str],
    start_page: int,
    page_size: int,
    limit: int,
    delay: float,
    timeout: float,
    seen_ids: Optional[set[str]] = None,
    skip_log_path: Optional[Path] = None,
    stop_on_detail_error: bool = False,
) -> Iterable[Dict[str, str]]:
    seen: set[str] = set(seen_ids or set())
    collected = 0
    page = start_page
    total = None

    while True:
        try:
            list_rows, total_count = fetch_list_page(
                session,
                page=page,
                page_size=page_size,
                timeout=timeout,
            )
        except Exception as exc:  # noqa: BLE001 - page window fallback.
            message = f"第 {page} 页批量读取失败：{exc}"
            print(message, file=sys.stderr)
            append_skip_log(skip_log_path, message)
            list_rows, total_count = fetch_list_page_one_by_one(
                session,
                page=page,
                page_size=page_size,
                timeout=timeout,
                known_total=total,
                skip_log_path=skip_log_path,
            )

        total = total_count if total is None else total
        if not list_rows:
            print(f"第 {page} 页没有可写入数据，停止抓取。", file=sys.stderr)
            break

        print(f"第 {page} 页：{len(list_rows)} 条，总数 {total_count}", file=sys.stderr)
        for item in list_rows:
            demand_id = clean_text(item.get("id"))
            if not demand_id or demand_id in seen:
                continue
            seen.add(demand_id)

            try:
                detail = fetch_detail(session, demand_id, timeout=timeout)
            except Exception as exc:  # noqa: BLE001 - keep crawling when a single detail endpoint is broken.
                if stop_on_detail_error:
                    message = f"需求 {demand_id} 详情读取失败，已暂停；重新运行脚本会从断点继续：{exc}"
                    print(message, file=sys.stderr)
                    append_skip_log(skip_log_path, message)
                    return
                message = f"需求 {demand_id} 详情读取失败，已用列表页信息占位并继续：{exc}"
                print(message, file=sys.stderr)
                append_skip_log(skip_log_path, message)
                detail = item

            yield extract_row(detail, area_map)
            collected += 1

            if limit and collected >= limit:
                return
            if delay > 0:
                time.sleep(delay)

        if total_count and page * page_size >= total_count:
            break
        page += 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="爬取 jstec.com.cn 技术需求列表与详情信息")
    parser.add_argument("--output", default="jstec_demands.xlsx", help="输出路径，支持 .xlsx 或 .csv")
    parser.add_argument("--jsonl", default="", help="可选：同时输出 JSONL 文件")
    parser.add_argument("--limit", type=int, default=0, help="最多抓取多少条；0 表示不限制")
    parser.add_argument("--start-page", type=int, default=1, help="从第几页开始")
    parser.add_argument("--page-size", type=int, default=10, help="每页条数；网站当前接口建议保持 10")
    parser.add_argument("--delay", type=float, default=0.3, help="每条详情请求后的等待秒数")
    parser.add_argument("--timeout", type=float, default=20, help="单次请求超时时间，秒")
    parser.add_argument("--login", action="store_true", help="登录后抓取；会提示输入用户名、密码和验证码")
    parser.add_argument("--username", default="", help="JSTEC 用户名/手机号；也可用环境变量 JSTEC_USERNAME")
    parser.add_argument("--password", default="", help="JSTEC 密码；也可用环境变量 JSTEC_PASSWORD")
    parser.add_argument("--verify-code", default="", help="验证码；不提供时会保存验证码图片并提示输入")
    parser.add_argument("--captcha-path", default="jstec_captcha.png", help="验证码图片保存路径")
    parser.add_argument("--checkpoint", default="", help="断点文件路径；默认与输出 Excel 同名")
    parser.add_argument("--fresh", action="store_true", help="重新开始抓取，忽略并清空已有断点")
    parser.add_argument("--backfill-intended-price", action="store_true", help="为旧断点记录回补“意向投入”；记录多时会花一些时间")
    parser.add_argument(
        "--stop-on-detail-error",
        action="store_true",
        help="遇到单条详情接口失败时暂停；默认会记录日志、写入列表页信息并继续",
    )
    args = parser.parse_args()
    if not args.username:
        args.username = os.environ.get("JSTEC_USERNAME", "")
    if not args.password:
        args.password = os.environ.get("JSTEC_PASSWORD", "")
    return args


def default_checkpoint_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}.checkpoint.jsonl")


def row_demand_id(row: Dict[str, str]) -> str:
    return clean_text(row.get("需求ID"))


def load_checkpoint(checkpoint_path: Path) -> list[Dict[str, str]]:
    if not checkpoint_path.exists():
        return []

    rows: list[Dict[str, str]] = []
    seen: set[str] = set()
    with checkpoint_path.open("r", encoding="utf-8") as file:
        for line_no, line in enumerate(file, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                print(f"断点文件第 {line_no} 行格式异常，已忽略。", file=sys.stderr)
                continue
            if not isinstance(row, dict):
                continue
            demand_id = row_demand_id(row)
            if demand_id and demand_id in seen:
                continue
            if demand_id:
                seen.add(demand_id)
            rows.append({field: clean_text(row.get(field)) for field in CSV_FIELDS})

    print(f"已读取断点 {checkpoint_path}：{len(rows)} 条。", file=sys.stderr)
    return rows


def rewrite_checkpoint(checkpoint_path: Path, rows: Iterable[Dict[str, str]]) -> None:
    tmp_path = checkpoint_path.with_name(f"{checkpoint_path.name}.tmp")
    with tmp_path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")
    tmp_path.replace(checkpoint_path)


def backfill_intended_price_for_checkpoint(
    session: requests.Session,
    rows: list[Dict[str, str]],
    checkpoint_path: Path,
    *,
    timeout: float,
    page_size: int,
    skip_log_path: Optional[Path] = None,
) -> None:
    missing_by_id = {row_demand_id(row): row for row in rows if row_demand_id(row) and not row.get("意向投入")}
    missing_rows = list(missing_by_id.values())
    if not missing_rows:
        return

    print(f"旧断点缺少“意向投入”，正在回补 {len(missing_rows)} 条。", file=sys.stderr)
    changed = False
    filled = 0
    page = 1
    total = None

    while missing_by_id:
        try:
            list_rows, total_count = fetch_list_page(
                session,
                page=page,
                page_size=page_size,
                timeout=timeout,
            )
        except Exception as exc:  # noqa: BLE001 - use the existing single-row fallback.
            message = f"回补意向投入时，第 {page} 页批量读取失败：{exc}"
            print(message, file=sys.stderr)
            append_skip_log(skip_log_path, message)
            list_rows, total_count = fetch_list_page_one_by_one(
                session,
                page=page,
                page_size=page_size,
                timeout=timeout,
                known_total=total,
                skip_log_path=skip_log_path,
            )

        if total is None and total_count:
            total = total_count
        if not list_rows:
            break

        for item in list_rows:
            demand_id = clean_text(item.get("id"))
            row = missing_by_id.pop(demand_id, None)
            if row is None:
                continue
            row["意向投入"] = format_intended_price(item.get("intendedPrice"))
            filled += 1
            changed = True

        if changed and (filled % 50 == 0 or not missing_by_id):
            rewrite_checkpoint(checkpoint_path, rows)
            print(f"已回补意向投入 {filled}/{len(missing_rows)} 条。", file=sys.stderr)

        if total_count and page * page_size >= total_count:
            break
        page += 1

    if missing_by_id:
        print(f"仍有 {len(missing_by_id)} 条未能从列表回补，改用详情接口补齐。", file=sys.stderr)
    for index, (demand_id, row) in enumerate(list(missing_by_id.items()), start=1):
        try:
            detail = fetch_detail(session, demand_id, timeout=timeout)
        except Exception as exc:  # noqa: BLE001 - keep the old row and continue.
            message = f"需求 {demand_id} 意向投入回补失败，已保留为空：{exc}"
            print(message, file=sys.stderr)
            append_skip_log(skip_log_path, message)
            continue
        row["意向投入"] = format_intended_price(detail.get("intendedPrice"))
        filled += 1
        changed = True
        if index % 20 == 0 or index == len(missing_by_id):
            rewrite_checkpoint(checkpoint_path, rows)
            print(f"已回补意向投入 {filled}/{len(missing_rows)} 条。", file=sys.stderr)

    if changed:
        rewrite_checkpoint(checkpoint_path, rows)
        print(f"已更新断点文件：{checkpoint_path}", file=sys.stderr)


def iter_with_checkpoint(
    rows: Iterable[Dict[str, str]],
    checkpoint_path: Path,
) -> Iterable[Dict[str, str]]:
    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    with checkpoint_path.open("a", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")
            file.flush()
            os.fsync(file.fileno())
            yield row


def prepend_rows(
    existing_rows: Iterable[Dict[str, str]],
    new_rows: Iterable[Dict[str, str]],
) -> Iterable[Dict[str, str]]:
    for row in existing_rows:
        yield row
    for row in new_rows:
        yield row


def write_xlsx(rows: Iterable[Dict[str, str]], output_path: Path, jsonl_path: Optional[Path]) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise RuntimeError("写入 .xlsx 需要安装 openpyxl：pip install openpyxl") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "技术需求"
    sheet.append(CSV_FIELDS)

    jsonl_file = None
    count = 0
    try:
        if jsonl_path:
            jsonl_file = jsonl_path.open("w", encoding="utf-8")

        for row in rows:
            sheet.append([row.get(field, "") for field in CSV_FIELDS])
            if jsonl_file:
                jsonl_file.write(json.dumps(row, ensure_ascii=False) + "\n")
            count += 1
            print(f"已写入 {count} 条：{row['需求名称']}", file=sys.stderr)
    finally:
        if jsonl_file:
            jsonl_file.close()

    header_font = Font(bold=True)
    wrap_fields = {"需求详情"}
    text_fields = {"需求编号", "联系方式", "需求ID"}
    widths = {
        "需求名称": 30,
        "需求编号": 22,
        "合作方式": 14,
        "意向投入": 14,
        "联系方式": 18,
        "发布者": 18,
        "需求详情": 70,
        "技术领域": 36,
        "需求类型": 18,
        "所在地区": 28,
        "需求ID": 24,
        "详情页链接": 60,
    }

    for col_index, field in enumerate(CSV_FIELDS, start=1):
        column_letter = sheet.cell(row=1, column=col_index).column_letter
        sheet.column_dimensions[column_letter].width = widths.get(field, 18)
        sheet.cell(row=1, column=col_index).font = header_font
        for cell in sheet[column_letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=field in wrap_fields)
            if field in text_fields:
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value)

    sheet.freeze_panes = "A2"
    workbook.save(output_path)
    return count


def write_csv(rows: Iterable[Dict[str, str]], output_path: Path, jsonl_path: Optional[Path]) -> int:
    count = 0
    jsonl_file = None
    try:
        if jsonl_path:
            jsonl_file = jsonl_path.open("w", encoding="utf-8")

        with output_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=CSV_FIELDS, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
                if jsonl_file:
                    jsonl_file.write(json.dumps(row, ensure_ascii=False) + "\n")
                count += 1
                print(f"已写入 {count} 条：{row['需求名称']}", file=sys.stderr)
    finally:
        if jsonl_file:
            jsonl_file.close()
    return count


def main() -> int:
    args = parse_args()
    output_path = Path(args.output)
    jsonl_path = Path(args.jsonl) if args.jsonl else None
    checkpoint_path = Path(args.checkpoint) if args.checkpoint else default_checkpoint_path(output_path)
    skip_log_path = output_path.with_name(f"{output_path.stem}.skipped.log")

    session = build_session()
    login_if_requested(session, args)
    area_map = fetch_area_code_map(session, args.timeout)
    if area_map:
        print(f"已读取 {len(area_map)} 个地区代码", file=sys.stderr)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if jsonl_path:
        jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)

    if args.fresh and checkpoint_path.exists():
        checkpoint_path.unlink()
        print(f"已清空旧断点：{checkpoint_path}", file=sys.stderr)
    if args.fresh and skip_log_path.exists():
        skip_log_path.unlink()
        print(f"已清空旧跳过日志：{skip_log_path}", file=sys.stderr)

    checkpoint_rows = load_checkpoint(checkpoint_path)
    if args.backfill_intended_price:
        backfill_intended_price_for_checkpoint(
            session,
            checkpoint_rows,
            checkpoint_path,
            timeout=args.timeout,
            page_size=args.page_size,
            skip_log_path=skip_log_path,
        )
    elif any(row_demand_id(row) and not row.get("意向投入") for row in checkpoint_rows):
        print(
            "提示：旧断点中有记录缺少“意向投入”。需要补齐时，可加参数 "
            "--backfill-intended-price 单独回补。",
            file=sys.stderr,
        )
    seen_ids = {row_demand_id(row) for row in checkpoint_rows if row_demand_id(row)}
    start_page = args.start_page
    if checkpoint_rows and args.start_page == 1:
        start_page = max(1, len(checkpoint_rows) // args.page_size + 1)
        print(f"将从第 {start_page} 页附近继续，并自动跳过已写入的需求。", file=sys.stderr)

    if args.limit and len(checkpoint_rows) >= args.limit:
        new_rows: Iterable[Dict[str, str]] = iter(())
        print(f"断点中已有 {len(checkpoint_rows)} 条，已达到 --limit {args.limit}。", file=sys.stderr)
    else:
        remaining_limit = args.limit - len(checkpoint_rows) if args.limit else 0
        new_rows = iter_demands(
            session,
            area_map=area_map,
            start_page=start_page,
            page_size=args.page_size,
            limit=remaining_limit,
            delay=args.delay,
            timeout=args.timeout,
            seen_ids=seen_ids,
            skip_log_path=skip_log_path,
            stop_on_detail_error=args.stop_on_detail_error,
        )
        new_rows = iter_with_checkpoint(new_rows, checkpoint_path)

    rows = prepend_rows(checkpoint_rows, new_rows)

    suffix = output_path.suffix.lower()
    if suffix == ".xlsx":
        count = write_xlsx(rows, output_path, jsonl_path)
    elif suffix == ".csv":
        count = write_csv(rows, output_path, jsonl_path)
        print("提示：CSV 用 Excel 直接打开时仍可能自动转科学计数法，建议输出 .xlsx。", file=sys.stderr)
    else:
        raise ValueError("输出文件后缀只支持 .xlsx 或 .csv")

    print(f"完成：共写入 {count} 条，文件：{output_path}", file=sys.stderr)
    print(f"断点文件：{checkpoint_path}", file=sys.stderr)
    if jsonl_path:
        print(f"JSONL：{jsonl_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
