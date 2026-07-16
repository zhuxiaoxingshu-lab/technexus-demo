from __future__ import annotations

import argparse
import hashlib
import hmac
import io
import json
import math
import mimetypes
import os
import re
import secrets
import sqlite3
import threading
import time
import urllib.error
import urllib.request
import uuid
import webbrowser
from collections import Counter
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse


APP_DIR = Path(__file__).resolve().parent
BASE_DIR = APP_DIR.parent
STATIC_DIR = APP_DIR / "static"
DATA_DIR = BASE_DIR / "technexus_data"
DEMANDS_FILE = BASE_DIR / "jstec_demands.checkpoint.jsonl"
AI_CONFIG_FILE = BASE_DIR / "ai_config.json"
ADMIN_CONFIG_FILE = BASE_DIR / "admin_config.json"
DB_FILE = DATA_DIR / "technexus.db"
DATABASE_URL = os.getenv("TECHNEXUS_DATABASE_URL") or os.getenv("DATABASE_URL") or ""
DEMAND_ANALYSIS_VERSION = "demand-profile-v1"
SESSION_COOKIE = "technexus_admin_session"
SESSION_SECONDS = 8 * 60 * 60
AGREEMENT_VERSION = "TechNexus-2026-06-01-v2"
PUBLIC_RESPONSE_PROMISE = "平台将在 3 个工作日内完成初步审核或联系。"
PROGRESS_STEPS = [
    ("submitted", "已提交合作意向", "已提交合作意向"),
    ("reviewing", "平台审核中", "平台审核中"),
    ("called_result_owner", "打电话核实成果", "已电话核实成果"),
    ("contacted_demander", "联系需求方", "已联系需求方"),
    ("wechat_contact", "发微信", "已微信沟通"),
    ("meeting_scheduled", "约线上会议", "已约线上会议"),
    ("agreement_sent", "发送协议", "已发送中介协议"),
    ("deal_recorded", "记录成交金额", "已记录合作结果"),
]
INTENT_STATUSES = [
    "待审核",
    "已联系成果方",
    "已联系需求方",
    "撮合中",
    "已签中介协议",
    "合作成功",
    "合作失败",
]
MATCH_FOLLOWUP_STATUSES = [
    "待联系成果方",
    "已联系成果方",
    "待联系需求方",
    "已联系需求方",
    "双方沟通中",
    "已安排会议",
    "已发送材料",
    "已签约",
    "已成交",
    "暂停跟进",
    "不再跟进",
]

CONTACT_FIELDS = {"联系方式", "详情页链接"}
DEMAND_FIELDS = [
    "需求名称",
    "需求编号",
    "合作方式",
    "意向投入",
    "联系方式",
    "发布者",
    "需求详情",
    "技术领域",
    "需求类型",
    "所在地区",
    "需求ID",
    "详情页链接",
]
DEMAND_NAME_FIELD = DEMAND_FIELDS[0]
DEMAND_NO_FIELD = DEMAND_FIELDS[1]
DEMAND_COOP_FIELD = DEMAND_FIELDS[2]
DEMAND_PRICE_FIELD = DEMAND_FIELDS[3]
DEMAND_CONTACT_FIELD = DEMAND_FIELDS[4]
DEMAND_PUBLISHER_FIELD = DEMAND_FIELDS[5]
DEMAND_DETAIL_FIELD = DEMAND_FIELDS[6]
DEMAND_TECH_FIELD = DEMAND_FIELDS[7]
DEMAND_TYPE_FIELD = DEMAND_FIELDS[8]
DEMAND_REGION_FIELD = DEMAND_FIELDS[9]
DEMAND_ID_FIELD = DEMAND_FIELDS[10]
DEMAND_LINK_FIELD = DEMAND_FIELDS[11]

STOPWORDS = {
    "技术",
    "需求",
    "成果",
    "项目",
    "研发",
    "开发",
    "应用",
    "合作",
    "解决",
    "相关",
    "研究",
    "产业",
    "材料",
    "产品",
    "进行",
    "实现",
    "提高",
    "具有",
    "通过",
    "基于",
    "以及",
    "一种",
    "方向",
    "领域",
}


TECH_TAG_VOCAB = {
    "新材料": ["石墨烯", "陶瓷", "高分子", "复合材料", "膜材料", "碳材料", "金属材料", "纳米材料"],
    "半导体": ["半导体", "晶圆", "芯片", "封装", "先进封装", "光刻", "刻蚀", "沉积", "薄膜", "ALD", "CVD", "LPCVD", "SiC", "GaN"],
    "新能源": ["新能源", "锂电", "钠电", "储能", "光伏", "风电", "氢能", "燃料电池", "电解槽", "电池"],
    "生物医药": ["生物医药", "合成生物", "发酵", "酶法", "药物", "医疗器械", "诊断", "蛋白", "细胞"],
    "智能制造": ["智能制造", "工业自动化", "机器人", "机械臂", "传感器", "数控", "机床", "视觉检测", "工业软件"],
    "电子信息": ["电子信息", "射频", "光电子", "激光", "MEMS", "通信", "雷达"],
    "环保低碳": ["环保", "废水", "废气", "固废", "减排", "低碳", "节能", "循环利用", "光催化"],
    "高端装备": ["高端装备", "装备制造", "成套设备", "精密加工", "真空设备", "工艺装备"],
}

SCENE_TAG_VOCAB = {
    "半导体封装": ["半导体封装", "先进封装", "晶圆制造", "晶圆", "芯片封装", "热界面"],
    "储能器件": ["储能", "电池", "锂电", "钠电", "电极", "电芯", "电池包"],
    "能源装备": ["光伏", "风电", "氢能", "燃料电池", "能源器件", "逆变器"],
    "工业废水治理": ["废水", "污水", "染料废水", "工业废水", "吸附", "光催化"],
    "海工船舶": ["海洋工程", "船舶", "海工", "海上风电", "耐蚀"],
    "汽车零部件": ["汽车", "车规", "零部件", "新能源车", "汽车电子"],
    "医疗健康": ["医疗", "医药", "诊断", "临床", "健康"],
    "家居建材": ["家居", "建材", "定制家居", "涂层", "装饰材料"],
    "食品农业": ["农业", "育种", "食品", "菌菇", "种植", "饲料"],
    "工厂产线": ["产线", "工厂", "自动化", "包装机", "生产线", "制造现场"],
}

INDUSTRY_TAG_VOCAB = {
    "材料制备": ["材料制备", "合成", "复合材料", "配方", "改性", "制膜", "制粉"],
    "工艺开发": ["工艺开发", "工艺优化", "工艺定型", "工艺包", "小试", "中试", "放大"],
    "器件开发": ["器件", "组件", "模组", "单元", "电池组件", "封装器件"],
    "装备制造": ["装备", "设备", "机台", "反应器", "打包机", "产线设备"],
    "检测验证": ["检测", "测试", "表征", "验证", "可靠性", "评价"],
    "系统集成": ["系统集成", "控制系统", "管理系统", "算法集成", "软件平台"],
    "量产导入": ["量产", "产业化", "导入", "示范线", "工厂化", "批量"],
}

COOPERATION_TAG_VOCAB = {
    "技术转让": ["技术转让", "成果转让", "专利转让", "license"],
    "合作开发": ["合作开发", "联合开发", "协同开发", "共同开发"],
    "技术服务": ["技术服务", "委托开发", "技术咨询", "解决方案"],
    "委托研发": ["委托研发", "外包研发", "委托开发"],
}

MATURITY_LABELS = {
    1: "概念阶段",
    2: "实验室阶段",
    3: "小试阶段",
    4: "中试阶段",
    5: "量产阶段",
}

PROBLEM_MARKERS = [
    "解决",
    "突破",
    "问题",
    "瓶颈",
    "痛点",
    "缺陷",
    "不足",
    "难题",
    "提升",
    "提高",
    "降低",
    "改善",
    "优化",
    "替代",
]
ROUTE_MARKERS = [
    "采用",
    "基于",
    "利用",
    "通过",
    "制备",
    "合成",
    "改性",
    "设计",
    "构建",
    "集成",
    "开发",
    "研发",
    "工艺",
    "算法",
    "模型",
    "装备",
    "系统",
]
CONSTRAINT_MARKERS = [
    "必须",
    "要求",
    "不低于",
    "不高于",
    "大于",
    "小于",
    "兼容",
    "满足",
    "成本",
    "温度",
    "压力",
    "精度",
    "寿命",
    "效率",
    "强度",
    "可靠性",
]
EVIDENCE_MARKERS = [
    "样品",
    "样机",
    "中试",
    "量产",
    "检测",
    "测试",
    "验证",
    "客户",
    "案例",
    "专利",
    "论文",
    "报告",
    "TRL",
]
DELIVERABLE_MARKERS = [
    "样品",
    "样机",
    "配方",
    "工艺包",
    "设备",
    "装备",
    "系统",
    "软件",
    "算法",
    "产线",
    "中试线",
    "检测报告",
    "解决方案",
]
GENERIC_TECH_TERMS = STOPWORDS | {
    "新材料",
    "电子信息",
    "智能制造",
    "先进制造",
    "新能源",
    "技术服务",
    "关键技术",
    "技术研发",
    "合作开发",
    "产业化",
    "技术领域",
    "应用场景",
    "材料制备",
    "器件开发",
    "工艺开发",
    "稳定",
    "稳定性",
    "可靠",
    "可靠性",
    "性能",
    "高性能",
    "效率",
    "高效",
    "精准",
    "自动化",
    "智能化",
    "管理",
    "控制",
    "优化",
    "改进",
    "加工",
    "生产",
    "制造",
    "规模化",
    "批量化",
    "可重复",
    "重复性",
    "适用",
    "安全",
    "环保",
    "节能",
    "降低成本",
    "质量",
    "方案",
    "方法",
    "系统",
    "装置",
}
GENERIC_ANCHOR_FRAGMENTS = {
    "稳定",
    "可靠",
    "性能",
    "效率",
    "高效",
    "精准",
    "自动",
    "智能",
    "管理",
    "控制",
    "优化",
    "改进",
    "加工",
    "生产",
    "制造",
    "规模",
    "批量",
    "重复",
    "适用",
    "安全",
    "环保",
    "节能",
    "成本",
    "质量",
    "方案",
    "方法",
    "系统",
    "装置",
    "技术",
    "需求",
    "成果",
    "项目",
    "研发",
    "开发",
    "应用",
    "合作",
    "解决",
    "研究",
    "材料",
    "产品",
    "工艺",
}
GENERIC_ANCHOR_PATTERN = re.compile(
    "|".join(re.escape(fragment) for fragment in sorted(GENERIC_ANCHOR_FRAGMENTS, key=len, reverse=True))
)
ANCHOR_FUNCTION_PREFIXES = tuple("的了在为与和及或对将由用于可需能从以并其该此把向")
ANCHOR_FUNCTION_SUFFIXES = ("的", "了", "中", "内", "外", "等", "方面", "相关")
MEASUREMENT_PATTERN = re.compile(
    r"[^。；;，,\n]{0,24}?(?:≥|≤|>|<|±|不低于|不高于|达到|控制在)?\s*"
    r"\d+(?:\.\d+)?(?:\s*[-~～至]\s*\d+(?:\.\d+)?)?\s*"
    r"(?:%|℃|°C|K|MPa|GPa|kPa|Pa|HV|HRC|W/\(m[·.]\s*K\)|W/mK|"
    r"mW/\(m[·.]\s*K\)|mm|μm|um|nm|cm|m|kg|g|mg|L|mL|h|小时|天|次|批次|"
    r"rpm|Hz|kHz|MHz|GHz|V|kV|A|mA|W|kW|MW|Wh|kWh|Torr|目|级)"
    r"[^。；;，,\n]{0,18}",
    re.IGNORECASE,
)


def dedupe_keep_order(values: list[str], limit: int = 0) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
        if limit and len(result) >= limit:
            break
    return result


def contains_alias(text: str, alias: str) -> bool:
    text = clean_text(text)
    alias = clean_text(alias)
    if not text or not alias:
        return False
    lowered = text.lower()
    alias_lower = alias.lower()
    if alias_lower.isascii():
        return alias_lower in lowered
    return alias in text


def extract_vocab_tags(text: str, vocab: dict[str, list[str]], *, limit: int = 6) -> list[str]:
    text = clean_text(text)
    if not text:
        return []
    scored: list[tuple[int, int, str]] = []
    for label, aliases in vocab.items():
        score = 0
        for alias in [label, *aliases]:
            if contains_alias(text, alias):
                score += 2 if alias == label else 1
        if score > 0:
            scored.append((score, len(label), label))
    scored.sort(key=lambda item: (item[0], item[1], item[2]), reverse=True)
    return [label for _, _, label in scored[:limit]]


def maturity_label(text: str) -> str:
    level = maturity_level(text)
    if level is None:
        return ""
    return MATURITY_LABELS.get(level, "")


def extract_region_tokens(text: str) -> list[str]:
    text = clean_text(text)
    if not text:
        return []
    parts = re.split(r"[\/,，、\-\s]+", text)
    tokens = []
    for part in parts:
        token = clean_text(part)
        if len(token) >= 2:
            tokens.append(token)
    return dedupe_keep_order(tokens, 6)


def top_keywords(text: str, limit: int = 8) -> list[str]:
    tokens = tokenize(text)
    ranked = sorted(tokens.items(), key=lambda item: (item[1], len(item[0]), item[0]), reverse=True)
    return [token for token, _ in ranked if len(token) <= 12][:limit]


def overlap_tags(left: list[str], right: list[str], *, limit: int = 6) -> list[str]:
    if not left or not right:
        return []
    right_set = set(right)
    shared = [item for item in left if item in right_set]
    return dedupe_keep_order(shared, limit)


def normalize_tag_payload(payload: dict | None) -> dict:
    payload = payload or {}
    return {
        "tech_tags": dedupe_keep_order(payload.get("tech_tags") or [], 6),
        "scene_tags": dedupe_keep_order(payload.get("scene_tags") or [], 6),
        "industry_tags": dedupe_keep_order(payload.get("industry_tags") or [], 6),
        "cooperation_tags": dedupe_keep_order(payload.get("cooperation_tags") or [], 4),
        "keywords": dedupe_keep_order(payload.get("keywords") or [], 10),
        "region_tokens": dedupe_keep_order(payload.get("region_tokens") or [], 6),
        "maturity_label": clean_text(payload.get("maturity_label", "")),
    }


def compact_tag_payload(payload: dict | None) -> dict:
    normalized = normalize_tag_payload(payload)
    return {
        "技术标签": normalized["tech_tags"],
        "应用标签": normalized["scene_tags"],
        "产业标签": normalized["industry_tags"],
        "合作标签": normalized["cooperation_tags"],
        "地区标签": normalized["region_tokens"],
        "成熟度标签": normalized["maturity_label"],
        "关键词": normalized["keywords"][:8],
    }


def tags_to_text(payload: dict | None) -> str:
    normalized = compact_tag_payload(payload)
    parts: list[str] = []
    for value in normalized.values():
        if isinstance(value, list):
            parts.extend(value)
        elif value:
            parts.append(value)
    return " ".join(parts)


def alias_lookup(vocab: dict[str, list[str]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for label, aliases in vocab.items():
        mapping[str(label).strip().lower()] = label
        for alias in aliases:
            mapping[str(alias).strip().lower()] = label
    return mapping


TECH_ALIAS_LOOKUP = alias_lookup(TECH_TAG_VOCAB)
SCENE_ALIAS_LOOKUP = alias_lookup(SCENE_TAG_VOCAB)
INDUSTRY_ALIAS_LOOKUP = alias_lookup(INDUSTRY_TAG_VOCAB)
COOP_ALIAS_LOOKUP = alias_lookup(COOPERATION_TAG_VOCAB)


def normalize_vocab_values(values: object, lookup: dict[str, str], vocab: dict[str, list[str]], *, limit: int) -> list[str]:
    result: list[str] = []
    items = values if isinstance(values, list) else [values]
    for raw in items:
        text = clean_text(raw)
        if not text:
            continue
        mapped = lookup.get(text.lower())
        if not mapped:
            guessed = extract_vocab_tags(text, vocab, limit=1)
            mapped = guessed[0] if guessed else text
        result.append(mapped)
    return dedupe_keep_order(result, limit)


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def load_ai_config() -> dict:
    config = {
        "enabled": False,
        "provider": "openai-compatible",
        "base_url": "",
        "api_key": "",
        "model": "",
        "temperature": 0.2,
        "timeout": 50,
    }
    if AI_CONFIG_FILE.exists():
        try:
            file_config = json.loads(AI_CONFIG_FILE.read_text(encoding="utf-8-sig"))
            if isinstance(file_config, dict):
                config.update(file_config)
        except (OSError, json.JSONDecodeError):
            pass

    env_map = {
        "TECHNEXUS_AI_BASE_URL": "base_url",
        "TECHNEXUS_AI_API_KEY": "api_key",
        "TECHNEXUS_AI_MODEL": "model",
        "TECHNEXUS_AI_PROVIDER": "provider",
        "TECHNEXUS_AI_TIMEOUT": "timeout",
    }
    for env_name, key in env_map.items():
        if os.getenv(env_name):
            config[key] = os.getenv(env_name, "")
    if os.getenv("TECHNEXUS_AI_ENABLED"):
        config["enabled"] = os.getenv("TECHNEXUS_AI_ENABLED", "").strip().lower() in {"1", "true", "yes", "on"}
    return config


def ai_is_configured(config: dict) -> bool:
    return bool(config.get("enabled") and config.get("base_url") and config.get("api_key") and config.get("model"))


def ai_status(config: dict) -> dict:
    configured = ai_is_configured(config)
    return {
        "ai_configured": configured,
        "ai_provider": clean_text(config.get("provider", "")) if configured else "",
        "ai_model": clean_text(config.get("model", "")) if configured else "",
        "ai_mode": "AI精排" if configured else "本地规则",
    }


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clip(text: str, limit: int = 220) -> str:
    text = clean_text(text)
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def format_money(value: object) -> str:
    text = clean_text(value)
    if not text or text in {"0", "0.0", "面议"}:
        return "面议"
    number_text = re.sub(r"[^\d.]", "", text)
    if not number_text:
        return text
    try:
        number = float(number_text)
    except ValueError:
        return text
    if number <= 0:
        return "面议"
    if number >= 10000:
        amount = number / 10000
        if amount.is_integer():
            return f"{int(amount)}万"
        return f"{amount:.1f}万"
    if number.is_integer():
        return f"{int(number)}元"
    return f"{number:.0f}元"


def public_demand_detail(demand: dict) -> str:
    """Return public requirement text with demand-side identity data removed."""
    detail = clean_text(demand.get("需求详情", ""))
    if not detail:
        return ""
    for hidden_value in (demand.get("发布者", ""), demand.get("联系方式", "")):
        hidden_text = clean_text(hidden_value)
        if hidden_text:
            detail = detail.replace(hidden_text, "（信息已隐藏）")
    detail = re.sub(r"(?<!\d)1[3-9]\d{9}(?!\d)", "（手机号已隐藏）", detail)
    detail = re.sub(r"(?<!\d)(?:0\d{2,3}[-－— ]?)?\d{7,8}(?!\d)", "（联系电话已隐藏）", detail)
    detail = re.sub(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", "（邮箱已隐藏）", detail)
    detail = re.sub(r"(?:联系人|联络人|负责人)\s*[:：]?\s*[\u4e00-\u9fff·]{2,10}", "联系人（姓名已隐藏）", detail)
    detail = re.sub(r"(?:微信号?|QQ)\s*[:：]?\s*[A-Za-z0-9_-]{5,}", "社交账号（已隐藏）", detail, flags=re.IGNORECASE)
    return detail


def extract_ascii_tokens(text: str) -> list[str]:
    return re.findall(r"[a-zA-Z0-9][a-zA-Z0-9.+#/-]{1,}", text.lower())


def extract_cjk_ngrams(text: str) -> list[str]:
    grams: list[str] = []
    for seq in re.findall(r"[\u4e00-\u9fff]+", text):
        if 2 <= len(seq) <= 6 and seq not in STOPWORDS:
            grams.append(seq)
        for size in (2, 3):
            if len(seq) >= size:
                grams.extend(seq[i : i + size] for i in range(len(seq) - size + 1))
    return grams


def tokenize(text: str) -> Counter:
    text = clean_text(text)
    tokens = extract_ascii_tokens(text) + extract_cjk_ngrams(text)
    return Counter(t for t in tokens if t and t not in STOPWORDS and len(t) >= 2)


def cosine(left: Counter, right: Counter) -> float:
    if not left or not right:
        return 0.0
    dot = sum(count * right.get(token, 0) for token, count in left.items())
    if dot <= 0:
        return 0.0
    left_norm = math.sqrt(sum(count * count for count in left.values()))
    right_norm = math.sqrt(sum(count * count for count in right.values()))
    if not left_norm or not right_norm:
        return 0.0
    return dot / (left_norm * right_norm)


def score_from_similarity(user_text: str, demand_text: str, *, baseline: int = 38, scale: float = 185.0) -> int:
    user_tokens = tokenize(user_text)
    if not user_tokens:
        return baseline
    demand_tokens = tokenize(demand_text)
    raw = cosine(user_tokens, demand_tokens)
    return max(0, min(100, round(raw * scale)))


def maturity_level(text: str) -> int | None:
    text = clean_text(text)
    if not text:
        return None
    mapping = [
        (5, ["量产", "产业化", "规模化", "已有客户", "批量"]),
        (4, ["中试", "示范线", "试生产", "工程化"]),
        (3, ["小试", "样品", "样机", "试制", "验证"]),
        (2, ["实验室", "原理样机", "概念验证", "TRL 3", "TRL3", "TRL 4", "TRL4"]),
        (1, ["概念", "方案", "论文"]),
    ]
    for level, words in mapping:
        if any(word in text for word in words):
            return level
    trl = re.search(r"TRL\s*([1-9])", text, re.IGNORECASE)
    if trl:
        value = int(trl.group(1))
        if value >= 8:
            return 5
        if value >= 6:
            return 4
        if value >= 4:
            return 3
        return 2
    return 3


def maturity_score(user_maturity: str, demand_detail: str) -> int:
    user_level = maturity_level(user_maturity)
    demand_level = maturity_level(demand_detail)
    if user_level is None:
        return 70
    if demand_level is None:
        demand_level = 3
    return max(35, 100 - abs(user_level - demand_level) * 18)


def shared_keywords(user_text: str, demand_text: str, limit: int = 6) -> list[str]:
    user_tokens = tokenize(user_text)
    demand_tokens = tokenize(demand_text)
    overlap = []
    for token in user_tokens:
        if token in demand_tokens and token not in STOPWORDS:
            overlap.append((len(token), user_tokens[token] + demand_tokens[token], token))
    overlap.sort(reverse=True)
    result: list[str] = []
    for _, _, token in overlap:
        if token not in result and len(token) <= 10:
            result.append(token)
        if len(result) >= limit:
            break
    return result


def split_technical_sentences(text: object) -> list[str]:
    cleaned = clean_text(text)
    if not cleaned:
        return []
    return dedupe_keep_order(
        [part.strip(" ：:；;，,。") for part in re.split(r"(?<=[。；;！？!?])|\n+", cleaned) if clean_text(part)],
        24,
    )


def sentences_with_markers(text: object, markers: list[str], *, limit: int = 5) -> list[str]:
    sentences = split_technical_sentences(text)
    matched = [sentence for sentence in sentences if any(marker.lower() in sentence.lower() for marker in markers)]
    return dedupe_keep_order(matched, limit)


def extract_measurements(text: object, *, limit: int = 8) -> list[str]:
    return dedupe_keep_order([clean_text(match.group(0)) for match in MEASUREMENT_PATTERN.finditer(clean_text(text))], limit)


DOMAIN_ANCHOR_TERMS = {
    term.lower()
    for vocabulary in (TECH_TAG_VOCAB, SCENE_TAG_VOCAB, INDUSTRY_TAG_VOCAB)
    for values in vocabulary.values()
    for term in values
    if len(term) >= 2 and term not in GENERIC_TECH_TERMS
}
DOMAIN_ANCHOR_TERMS.update(
    {
        "导热",
        "高导热",
        "散热",
        "界面结合",
        "热管理",
        "抗热冲击",
        "冷热循环",
        "粘结",
        "导电",
        "绝缘",
        "耐腐蚀",
        "耐磨",
        "阻燃",
        "催化",
        "吸附",
        "分离",
        "过滤",
        "清洗",
        "振动",
        "高频振动",
        "超声",
        "焊接",
        "切削",
        "涂层",
        "封装",
        "覆铜",
        "烧结",
        "改性",
        "发酵",
        "酶法",
        "识别",
        "检测",
        "定位",
        "预测",
    }
)
FUNCTIONAL_ANCHOR_TERMS = {
    "导热",
    "高导热",
    "散热",
    "热管理",
    "传热",
    "冷却",
    "降温",
    "隔热",
    "导电",
    "绝缘",
    "电磁屏蔽",
    "耐腐蚀",
    "防腐",
    "耐磨",
    "阻燃",
    "抗菌",
    "消毒",
    "除尘",
    "脱硫",
    "脱硝",
    "减振",
    "降噪",
    "密封",
    "防水",
    "润滑",
    "增强",
    "增韧",
    "催化",
    "吸附",
    "分离",
    "过滤",
    "清洗",
    "振动",
    "高频振动",
    "超声",
    "焊接",
    "切削",
    "发酵",
    "酶法",
    "识别",
    "检测",
    "定位",
    "预测",
}


def is_generic_technical_term(term: str) -> bool:
    normalized = clean_text(term).lower()
    if not normalized or normalized in GENERIC_TECH_TERMS:
        return True
    if normalized in DOMAIN_ANCHOR_TERMS:
        return False
    if normalized.startswith(ANCHOR_FUNCTION_PREFIXES) or normalized.endswith(ANCHOR_FUNCTION_SUFFIXES):
        return True
    return bool(GENERIC_ANCHOR_PATTERN.search(normalized))


def extract_technical_anchors(text: object) -> set[str]:
    cleaned = clean_text(text).lower()
    if not cleaned:
        return set()
    anchors = {term for term in DOMAIN_ANCHOR_TERMS if term in cleaned}
    anchors.update(extract_ascii_tokens(cleaned))
    for sequence in re.findall(r"[\u4e00-\u9fff]+", cleaned):
        if 2 <= len(sequence) <= 8 and not is_generic_technical_term(sequence):
            anchors.add(sequence)
        for size in (3, 4):
            if len(sequence) < size:
                continue
            for index in range(len(sequence) - size + 1):
                term = sequence[index : index + size]
                if not is_generic_technical_term(term):
                    anchors.add(term)
    return anchors


def extract_functional_anchors(text: object) -> set[str]:
    cleaned = clean_text(text).lower()
    return {term for term in FUNCTIONAL_ANCHOR_TERMS if term in cleaned}


def rank_technical_anchors(shared: set[str], limit: int = 8) -> list[str]:
    ranked = sorted(
        shared,
        key=lambda term: (
            term in DOMAIN_ANCHOR_TERMS,
            len(term),
            term,
        ),
        reverse=True,
    )
    result: list[str] = []
    for term in ranked:
        if any(term != kept and term in kept for kept in result):
            continue
        result.append(term)
        if len(result) >= limit:
            break
    return result


def shared_technical_anchors(left: object, right: object, limit: int = 8) -> list[str]:
    return rank_technical_anchors(extract_technical_anchors(left) & extract_technical_anchors(right), limit)


def display_technical_anchors(values: list[str], limit: int = 8) -> list[str]:
    unique = set(values)
    preferred = {term for term in unique if term in DOMAIN_ANCHOR_TERMS}
    fallback = {term for term in unique if len(term) >= 4 and not is_generic_technical_term(term)}
    return rank_technical_anchors(preferred or fallback or unique, limit)


def specific_keywords(text: object, *, limit: int = 12) -> list[str]:
    values = []
    for keyword in top_keywords(clean_text(text), limit * 3):
        if is_generic_technical_term(keyword) or len(keyword) < 2:
            continue
        if any(keyword != kept and keyword in kept for kept in values):
            continue
        values.append(keyword)
        if len(values) >= limit:
            break
    return values


def list_values(value: object) -> list[object]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    text = clean_text(value)
    return [text] if text else []


def normalize_technical_profile(payload: dict | None) -> dict:
    payload = payload or {}
    return {
        "target": clip(payload.get("target", ""), 220),
        "core_problem": clip(payload.get("core_problem", ""), 420),
        "required_functions": dedupe_keep_order(list_values(payload.get("required_functions")), 6),
        "technical_route": clip(payload.get("technical_route", ""), 420),
        "indicators": dedupe_keep_order(list_values(payload.get("indicators")), 8),
        "constraints": dedupe_keep_order(list_values(payload.get("constraints")), 6),
        "application_object": clip(payload.get("application_object", ""), 220),
        "deliverables": dedupe_keep_order(list_values(payload.get("deliverables")), 6),
        "evidence": dedupe_keep_order(list_values(payload.get("evidence")), 6),
        "maturity": clean_text(payload.get("maturity", "")),
        "target_terms": dedupe_keep_order(list_values(payload.get("target_terms")), 12),
        "problem_terms": dedupe_keep_order(list_values(payload.get("problem_terms")), 12),
        "route_terms": dedupe_keep_order(list_values(payload.get("route_terms")), 12),
        "indicator_terms": dedupe_keep_order(list_values(payload.get("indicator_terms")), 12),
    }


def build_demand_technical_profile(demand: dict) -> dict:
    title = clean_text(demand.get(DEMAND_NAME_FIELD, ""))
    detail = clean_text(demand.get(DEMAND_DETAIL_FIELD, ""))
    problem_sentences = sentences_with_markers(detail, PROBLEM_MARKERS, limit=5)
    route_sentences = sentences_with_markers(detail, ROUTE_MARKERS, limit=5)
    constraint_sentences = sentences_with_markers(detail, CONSTRAINT_MARKERS, limit=5)
    deliverable_sentences = sentences_with_markers(detail, DELIVERABLE_MARKERS, limit=4)
    indicators = extract_measurements(detail)
    core_problem = clip(" ".join(problem_sentences) or detail, 420)
    technical_route = clip(" ".join(route_sentences) or detail, 420)
    application_object = clip(" ".join(split_technical_sentences(detail)[:2]), 220)
    deliverables = specific_keywords(" ".join(deliverable_sentences), limit=6)
    target_text = " ".join([title, application_object])
    return normalize_technical_profile(
        {
            "target": title,
            "core_problem": core_problem,
            "required_functions": specific_keywords(core_problem, limit=6),
            "technical_route": technical_route,
            "indicators": indicators,
            "constraints": constraint_sentences,
            "application_object": application_object,
            "deliverables": deliverables,
            "evidence": sentences_with_markers(detail, EVIDENCE_MARKERS, limit=6),
            "maturity": maturity_label(detail),
            "target_terms": specific_keywords(target_text, limit=12),
            "problem_terms": specific_keywords(core_problem, limit=12),
            "route_terms": specific_keywords(technical_route, limit=12),
            "indicator_terms": specific_keywords(clip(" ".join(indicators + constraint_sentences), 420), limit=12),
        }
    )


def demand_content_hash(demand: dict) -> str:
    """Return a stable fingerprint for fields that affect technical analysis."""
    payload = {
        field: clean_text(demand.get(field, ""))
        for field in (
            DEMAND_NAME_FIELD,
            DEMAND_DETAIL_FIELD,
            DEMAND_TECH_FIELD,
            DEMAND_TYPE_FIELD,
            DEMAND_COOP_FIELD,
            DEMAND_PRICE_FIELD,
            DEMAND_REGION_FIELD,
        )
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def build_submission_technical_profile(submission: dict) -> dict:
    title = clean_text(submission.get("title") or submission.get("achievement_name", ""))
    achievement_text = clean_text(submission.get("achievement_text") or submission.get("full_text", ""))
    summary = clean_text(submission.get("summary", ""))
    problem = clean_text(submission.get("problem", ""))
    advantages = clean_text(submission.get("advantages", ""))
    scene = clean_text(submission.get("application_scene", ""))
    explicit_route = clean_text(submission.get("technical_route", ""))
    explicit_indicators = clean_text(submission.get("indicators", ""))
    explicit_evidence = clean_text(submission.get("evidence", ""))
    extra = clean_text(submission.get("extra_note") or submission.get("attachment_note", ""))
    combined = " ".join(
        [
            title,
            achievement_text,
            summary,
            problem,
            explicit_route,
            explicit_indicators,
            explicit_evidence,
            advantages,
            scene,
            extra,
        ]
    )
    route_sentences = sentences_with_markers(
        " ".join([explicit_route, achievement_text, summary, advantages, extra]), ROUTE_MARKERS, limit=5
    )
    problem_text = problem or " ".join(
        sentences_with_markers(" ".join([achievement_text, summary]), PROBLEM_MARKERS, limit=5)
    )
    technical_route = clip(explicit_route or " ".join(route_sentences) or summary or achievement_text, 420)
    application_object = scene or clip(" ".join(split_technical_sentences(achievement_text)[:2]), 220)
    indicators = dedupe_keep_order(
        [*extract_measurements(explicit_indicators), *extract_measurements(combined)], 8
    )
    evidence = dedupe_keep_order(
        [explicit_evidence, *sentences_with_markers(combined, EVIDENCE_MARKERS, limit=6)], 6
    )
    return normalize_technical_profile(
        {
            "target": title,
            "core_problem": problem_text,
            "required_functions": specific_keywords(" ".join([problem_text, summary, achievement_text]), limit=6),
            "technical_route": technical_route,
            "indicators": indicators,
            "constraints": sentences_with_markers(combined, CONSTRAINT_MARKERS, limit=5),
            "application_object": application_object,
            "deliverables": specific_keywords(
                " ".join([clean_text(submission.get("cooperation", "")), clean_text(submission.get("maturity", "")), summary]),
                limit=6,
            ),
            "evidence": evidence,
            "maturity": maturity_label(clean_text(submission.get("maturity", ""))),
            "target_terms": specific_keywords(" ".join([title, application_object, achievement_text]), limit=12),
            "problem_terms": specific_keywords(" ".join([problem_text, summary, achievement_text]), limit=12),
            "route_terms": specific_keywords(" ".join([technical_route, advantages, achievement_text]), limit=12),
            "indicator_terms": specific_keywords(" ".join(indicators), limit=12),
        }
    )


def merge_technical_profiles(local_profile: dict, ai_profile: dict | None) -> dict:
    local = normalize_technical_profile(local_profile)
    ai = normalize_technical_profile(ai_profile)
    merged = dict(local)
    for field in ("target", "core_problem", "technical_route", "application_object", "maturity"):
        if ai.get(field):
            merged[field] = ai[field]
    for field in (
        "required_functions",
        "indicators",
        "constraints",
        "deliverables",
        "evidence",
        "target_terms",
        "problem_terms",
        "route_terms",
        "indicator_terms",
    ):
        merged[field] = dedupe_keep_order([*(ai.get(field) or []), *(local.get(field) or [])], 12)
    return normalize_technical_profile(merged)


def technical_profile_text(profile: dict, *fields: str) -> str:
    parts: list[str] = []
    for field in fields:
        value = profile.get(field)
        if isinstance(value, list):
            parts.extend(value)
        elif value:
            parts.append(str(value))
    return " ".join(parts)


def technical_similarity(left: str, right: str, *, scale: float = 235.0) -> int:
    if not clean_text(left) or not clean_text(right):
        return 0
    return max(0, min(100, round(cosine(tokenize(left), tokenize(right)) * scale)))


def anchored_technical_similarity(
    left: str,
    right: str,
    *,
    scale: float = 235.0,
    left_anchors: set[str] | None = None,
    right_anchors: set[str] | None = None,
) -> tuple[int, list[str]]:
    raw_score = technical_similarity(left, right, scale=scale)
    anchors = rank_technical_anchors(
        (left_anchors if left_anchors is not None else extract_technical_anchors(left))
        & (right_anchors if right_anchors is not None else extract_technical_anchors(right)),
        8,
    )
    if not anchors:
        return min(raw_score, 18), []
    longest = max(len(anchor) for anchor in anchors)
    anchor_strength = min(92, 34 + max(0, longest - 2) * 12 + max(0, len(anchors) - 1) * 7)
    score = round(raw_score * 0.28 + anchor_strength * 0.72)
    return max(0, min(100, score)), anchors


def submission_confidence(submission: dict, profile: dict, demand_profile: dict) -> int:
    score = 18
    weights = {
        "title": 8,
        "summary": 16,
        "problem": 14,
        "technical_route": 10,
        "indicators": 10,
        "evidence": 8,
        "application_scene": 8,
        "advantages": 8,
        "maturity": 7,
        "cooperation": 4,
        "extra_note": 4,
    }
    for field, weight in weights.items():
        if clean_text(submission.get(field, "")):
            score += weight
    if profile.get("indicators"):
        score += 8
    if profile.get("evidence"):
        score += 8
    if demand_profile.get("indicators"):
        score += 4
    if demand_profile.get("core_problem") and demand_profile.get("technical_route"):
        score += 5
    return max(20, min(95, score))


def ensure_demand_technical_profile(demand: dict) -> dict:
    existing = demand.get("_technical_profile")
    profile = normalize_technical_profile(existing) if existing else build_demand_technical_profile(demand)
    demand["_technical_profile"] = profile
    demand["_problem_tokens"] = tokenize(technical_profile_text(profile, "core_problem", "problem_terms"))
    demand["_target_tokens"] = tokenize(technical_profile_text(profile, "target", "application_object", "target_terms"))
    demand["_route_tokens"] = tokenize(technical_profile_text(profile, "technical_route", "route_terms"))
    demand["_indicator_tokens"] = tokenize(
        technical_profile_text(profile, "indicators", "constraints", "indicator_terms")
    )
    demand["_problem_anchors"] = extract_technical_anchors(
        technical_profile_text(profile, "core_problem", "required_functions", "problem_terms")
    )
    demand["_target_anchors"] = extract_technical_anchors(
        technical_profile_text(profile, "target", "application_object", "target_terms")
    )
    demand["_route_anchors"] = extract_technical_anchors(
        technical_profile_text(profile, "technical_route", "route_terms")
    )
    demand["_function_anchors"] = extract_functional_anchors(
        technical_profile_text(
            profile,
            "target",
            "core_problem",
            "required_functions",
            "technical_route",
            "application_object",
        )
    )
    return profile


class DemandStore:
    def __init__(self, path: Path):
        self.path = path
        self.demands: list[dict] = []
        self.loaded_at = ""
        self.source_version = ""
        self._last_refresh_check = 0.0

    def load(self) -> None:
        database_version = database_demands_version()
        database_demands = load_demands_from_database()
        if database_demands:
            self.demands = database_demands
            self.loaded_at = now_iso()
            self.source_version = database_version
            return

        self.demands = load_demands_from_file(self.path)
        self.loaded_at = now_iso()
        try:
            stat = self.path.stat()
            self.source_version = f"file:{stat.st_mtime_ns}:{stat.st_size}"
        except OSError:
            self.source_version = f"file:{self.loaded_at}"

    def reload(self) -> None:
        self.load()

    def refresh_if_changed(self, *, min_interval: float = 60.0) -> None:
        now = time.time()
        if now - self._last_refresh_check < min_interval:
            return
        self._last_refresh_check = now
        database_version = database_demands_version()
        if database_version and database_version != self.source_version:
            self.load()
            return
        if not database_version and self.path.exists():
            try:
                stat = self.path.stat()
                file_version = f"file:{stat.st_mtime_ns}:{stat.st_size}"
            except OSError:
                return
            if file_version != self.source_version:
                self.load()

    def stats(self) -> dict:
        self.refresh_if_changed()
        analysis = demand_analysis_stats()
        ai_profile_count = sum(
            int(item.get("count") or 0)
            for item in analysis.get("groups", [])
            if clean_text(item.get("source")) == "local+ai"
        )
        return {
            "demand_count": len(self.demands),
            "loaded_at": self.loaded_at,
            "demand_profile_count": int(analysis.get("total") or 0),
            "demand_ai_profile_count": ai_profile_count,
        }

    def by_id(self, demand_id: str) -> dict | None:
        demand_id = clean_text(demand_id)
        for demand in self.demands:
            if demand.get("需求ID") == demand_id:
                return demand
        return None

    def search(self, keyword: str = "", limit: int = 50) -> list[dict]:
        items, _ = self.search_page(keyword=keyword, offset=0, limit=limit)
        return items

    def search_page(self, keyword: str = "", offset: int = 0, limit: int = 24) -> tuple[list[dict], int]:
        self.refresh_if_changed()
        keyword = clean_text(keyword)
        offset = max(0, int(offset or 0))
        limit = max(1, min(60, int(limit or 24)))
        if not keyword:
            matched_rows = self.demands
        else:
            keyword_tokens = tokenize(keyword)
            scored = []
            for demand in self.demands:
                score = cosine(keyword_tokens, demand.get("_tokens", Counter()))
                if keyword in demand.get("_search_text", ""):
                    score += 0.35
                if score > 0:
                    scored.append((score, demand))
            scored.sort(key=lambda item: item[0], reverse=True)
            matched_rows = [demand for _, demand in scored]
        total = len(matched_rows)
        rows = matched_rows[offset : offset + limit]
        return [sanitize_demand(row, include_detail=True) for row in rows], total


def prepare_demand_row(item: dict) -> dict:
    cleaned = {clean_text(k): clean_text(v) for k, v in item.items()}
    cleaned["_structured_tags"] = extract_demand_tags(cleaned)
    tag_text = tags_to_text(cleaned["_structured_tags"])
    if not cleaned.get("需求ID") and not cleaned.get("需求名称"):
        return {}
    search_text = " ".join(
        cleaned.get(field, "")
        for field in ("需求名称", "技术领域", "需求类型", "所在地区", "需求详情", "合作方式", "发布者")
    )
    cleaned["_search_text"] = " ".join([search_text, tag_text]).strip()
    cleaned["_tokens"] = tokenize(cleaned["_search_text"])
    return cleaned


def load_demands_from_file(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"未找到需求库文件：{path}")
    demands: list[dict] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            cleaned = prepare_demand_row(item)
            if not cleaned:
                continue
            demands.append(cleaned)
    return demands


def sanitize_demand(demand: dict, *, include_detail: bool = False) -> dict:
    item = {
        "demand_id": demand.get("需求ID", ""),
        "name": demand.get("需求名称", ""),
        "demand_no": demand.get("需求编号", ""),
        "cooperation_mode": demand.get("合作方式", ""),
        "intended_price": format_money(demand.get("意向投入", "")),
        "tech_field": demand.get("技术领域", ""),
        "demand_type": demand.get("需求类型", ""),
        "region": demand.get("所在地区", ""),
        "publisher": demand.get("发布者", ""),
    }
    if include_detail:
        item["detail_summary"] = clip(demand.get("需求详情", ""), 260)
    return item


def public_demand_payload(item: dict) -> dict:
    """Remove demand-side identity and contact fields from user-facing responses."""
    hidden_fields = {"publisher", "contact", "source_url", "full_detail", "technical_profile"}
    return {key: value for key, value in item.items() if key not in hidden_fields}


def admin_demand_payload(demand: dict) -> dict:
    item = sanitize_demand(demand, include_detail=True)
    item.update(
        {
            "publisher": clean_text(demand.get("发布者")),
            "contact": clean_text(demand.get("联系方式")),
            "source_url": clean_text(demand.get("详情页链接")),
            "full_detail": clean_text(demand.get("需求详情")),
        }
    )
    return item


def build_user_text(data: dict, fields: tuple[str, ...]) -> str:
    return " ".join(clean_text(data.get(field, "")) for field in fields)


def extract_submission_tags_local(submission: dict) -> dict:
    tech_text = build_user_text(
        submission,
        (
            "tech_field",
            "title",
            "achievement_text",
            "summary",
            "technical_route",
            "advantages",
            "problem",
            "extra_note",
            "attachment_note",
        ),
    )
    scene_text = build_user_text(
        submission,
        ("application_scene", "achievement_text", "summary", "problem", "indicators", "extra_note", "attachment_note"),
    )
    industry_text = build_user_text(
        submission,
        (
            "tech_field",
            "application_scene",
            "achievement_text",
            "summary",
            "technical_route",
            "cooperation",
            "advantages",
            "extra_note",
            "attachment_note",
        ),
    )
    cooperation_text = build_user_text(
        submission,
        ("cooperation", "achievement_text", "summary", "evidence", "extra_note", "attachment_note"),
    )
    full_text = build_user_text(
        submission,
        (
            "title",
            "achievement_text",
            "tech_field",
            "application_scene",
            "summary",
            "technical_route",
            "indicators",
            "evidence",
            "advantages",
            "problem",
            "cooperation",
            "region",
            "extra_note",
            "attachment_note",
        ),
    )
    return normalize_tag_payload(
        {
            "tech_tags": extract_vocab_tags(tech_text, TECH_TAG_VOCAB, limit=6),
            "scene_tags": extract_vocab_tags(scene_text, SCENE_TAG_VOCAB, limit=6),
            "industry_tags": extract_vocab_tags(industry_text, INDUSTRY_TAG_VOCAB, limit=6),
            "cooperation_tags": extract_vocab_tags(cooperation_text, COOPERATION_TAG_VOCAB, limit=4),
            "keywords": top_keywords(full_text, 10),
            "region_tokens": extract_region_tokens(clean_text(submission.get("region", ""))),
            "maturity_label": maturity_label(clean_text(submission.get("maturity", ""))),
        }
    )


def extract_demand_tags(demand: dict) -> dict:
    tech_text = " ".join([demand.get(DEMAND_NAME_FIELD, ""), demand.get(DEMAND_TECH_FIELD, ""), demand.get(DEMAND_DETAIL_FIELD, "")])
    scene_text = " ".join([demand.get(DEMAND_NAME_FIELD, ""), demand.get(DEMAND_DETAIL_FIELD, "")])
    industry_text = " ".join([demand.get(DEMAND_TECH_FIELD, ""), demand.get(DEMAND_TYPE_FIELD, ""), demand.get(DEMAND_COOP_FIELD, ""), demand.get(DEMAND_DETAIL_FIELD, "")])
    all_text = " ".join(
        [
            demand.get(DEMAND_NAME_FIELD, ""),
            demand.get(DEMAND_TECH_FIELD, ""),
            demand.get(DEMAND_TYPE_FIELD, ""),
            demand.get(DEMAND_REGION_FIELD, ""),
            demand.get(DEMAND_COOP_FIELD, ""),
            demand.get(DEMAND_DETAIL_FIELD, ""),
        ]
    )
    return normalize_tag_payload(
        {
            "tech_tags": extract_vocab_tags(tech_text, TECH_TAG_VOCAB, limit=6),
            "scene_tags": extract_vocab_tags(scene_text, SCENE_TAG_VOCAB, limit=6),
            "industry_tags": extract_vocab_tags(industry_text, INDUSTRY_TAG_VOCAB, limit=6),
            "cooperation_tags": extract_vocab_tags(demand.get(DEMAND_COOP_FIELD, ""), COOPERATION_TAG_VOCAB, limit=4),
            "keywords": top_keywords(all_text, 12),
            "region_tokens": extract_region_tokens(demand.get(DEMAND_REGION_FIELD, "")),
            "maturity_label": maturity_label(demand.get(DEMAND_DETAIL_FIELD, "")),
        }
    )


def merge_tag_profiles(*profiles: dict | None) -> dict:
    merged = normalize_tag_payload({})
    for profile in profiles:
        current = normalize_tag_payload(profile)
        for field in ("tech_tags", "scene_tags", "industry_tags", "cooperation_tags", "keywords", "region_tokens"):
            max_limit = 10 if field == "keywords" else 6
            merged[field] = dedupe_keep_order([*merged[field], *current[field]], max_limit)
        if not merged["maturity_label"] and current["maturity_label"]:
            merged["maturity_label"] = current["maturity_label"]
    return merged


def maturity_distance_score(left_label: str, right_label: str) -> int:
    reverse = {value: key for key, value in MATURITY_LABELS.items()}
    left = reverse.get(clean_text(left_label))
    right = reverse.get(clean_text(right_label))
    if not left or not right:
        return 0
    return max(0, 16 - abs(left - right) * 6)


def recall_score_demand(
    submission: dict,
    demand: dict,
    tags: dict,
    capability_profile: dict | None = None,
    profile_tokens: dict | None = None,
) -> float:
    profile = normalize_technical_profile(capability_profile or build_submission_technical_profile(submission))
    profile_tokens = profile_tokens or {
        "problem": tokenize(technical_profile_text(profile, "core_problem", "problem_terms", "required_functions")),
        "target": tokenize(technical_profile_text(profile, "target", "application_object", "target_terms")),
        "route": tokenize(technical_profile_text(profile, "technical_route", "route_terms")),
        "indicator": tokenize(technical_profile_text(profile, "indicators", "constraints", "indicator_terms")),
        "problem_anchors": extract_technical_anchors(
            technical_profile_text(profile, "core_problem", "problem_terms", "required_functions")
        ),
        "target_anchors": extract_technical_anchors(
            technical_profile_text(profile, "target", "application_object", "target_terms")
        ),
        "route_anchors": extract_technical_anchors(
            technical_profile_text(profile, "technical_route", "route_terms")
        ),
        "function_anchors": extract_functional_anchors(
            technical_profile_text(
                profile,
                "target",
                "core_problem",
                "required_functions",
                "technical_route",
                "application_object",
            )
        ),
        "all": tokenize(
            build_user_text(
                submission,
                (
                    "title",
                    "application_scene",
                    "summary",
                    "technical_route",
                    "indicators",
                    "evidence",
                    "advantages",
                    "problem",
                    "extra_note",
                    "attachment_note",
                ),
            )
        ),
    }
    problem_tokens = profile_tokens.get("problem", Counter())
    target_tokens = profile_tokens.get("target", Counter())
    route_tokens = profile_tokens.get("route", Counter())
    indicator_tokens = profile_tokens.get("indicator", Counter())
    lexical = cosine(profile_tokens.get("all", Counter()), demand.get("_tokens", Counter()))
    problem_similarity = cosine(problem_tokens, demand.get("_problem_tokens", Counter())) if problem_tokens else 0.0
    target_similarity = cosine(target_tokens, demand.get("_target_tokens", Counter())) if target_tokens else 0.0
    route_similarity = cosine(route_tokens, demand.get("_route_tokens", Counter())) if route_tokens else 0.0
    indicator_similarity = cosine(indicator_tokens, demand.get("_indicator_tokens", Counter())) if indicator_tokens else 0.0
    demand_profile = normalize_technical_profile(demand.get("_technical_profile"))
    problem_anchors = rank_technical_anchors(
        set(profile_tokens.get("problem_anchors", set())) & set(demand.get("_problem_anchors", set())),
        6,
    )
    target_anchors = rank_technical_anchors(
        set(profile_tokens.get("target_anchors", set())) & set(demand.get("_target_anchors", set())),
        6,
    )
    route_anchors = rank_technical_anchors(
        set(profile_tokens.get("route_anchors", set())) & set(demand.get("_route_anchors", set())),
        6,
    )
    function_anchors = rank_technical_anchors(
        set(profile_tokens.get("function_anchors", set())) & set(demand.get("_function_anchors", set())),
        6,
    )
    demand_tags = normalize_tag_payload(demand.get("_structured_tags"))
    tech_overlap = overlap_tags(tags.get("tech_tags", []), demand_tags.get("tech_tags", []))
    scene_overlap = overlap_tags(tags.get("scene_tags", []), demand_tags.get("scene_tags", []))
    industry_overlap = overlap_tags(tags.get("industry_tags", []), demand_tags.get("industry_tags", []))
    keyword_overlap = overlap_tags(tags.get("keywords", []), demand_tags.get("keywords", []), limit=8)

    # Industry and broad tags only help recall. Concrete problem, target and route dominate.
    score = (
        problem_similarity * 1.35
        + target_similarity * 1.15
        + route_similarity * 1.10
        + indicator_similarity * 0.45
        + lexical * 0.30
    )
    score += min(0.16, len(tech_overlap) * 0.04)
    score += min(0.12, len(scene_overlap) * 0.04)
    score += min(0.08, len(industry_overlap) * 0.025)
    score += min(0.12, len(keyword_overlap) * 0.025)
    score += min(0.45, len(target_anchors) * 0.16)
    score += min(0.36, len(problem_anchors) * 0.12)
    score += min(0.40, len(route_anchors) * 0.14)
    score += min(0.50, len(function_anchors) * 0.18)
    if not target_anchors and not problem_anchors and not route_anchors:
        score *= 0.22
    if profile.get("target") and clean_text(profile["target"]) in demand.get("_search_text", ""):
        score += 0.18
    return score


def indicator_fit_score(profile: dict, demand_profile: dict) -> tuple[int, list[str], list[str]]:
    user_indicators = profile.get("indicators") or []
    demand_indicators = demand_profile.get("indicators") or []
    if not demand_indicators:
        return (62 if user_indicators else 52), [], []
    if not user_indicators:
        return 32, [], demand_indicators[:4]
    similarity = technical_similarity(" ".join(user_indicators), " ".join(demand_indicators), scale=250)
    shared = shared_keywords(" ".join(user_indicators), " ".join(demand_indicators), 5)
    score = max(20, min(100, similarity + min(20, len(shared) * 7)))
    return score, shared, [item for item in demand_indicators if not any(term in item for term in shared)][:4]


def classify_match_type(
    problem_score: int,
    target_score: int,
    route_score: int,
    tech_overlap: list[str],
    problem_anchors: list[str],
    target_anchors: list[str],
    route_anchors: list[str],
    *,
    has_explicit_route: bool = False,
    has_explicit_problem: bool = False,
) -> str:
    if (
        has_explicit_problem
        and target_anchors
        and problem_anchors
        and target_score >= 58
        and problem_score >= 50
    ):
        return "直接解决"
    if not target_anchors and route_anchors and (has_explicit_route or len(route_anchors) >= 2) and route_score >= 48:
        return "技术迁移"
    if target_anchors and route_anchors and route_score >= 52:
        return "关键组件"
    if tech_overlap and not (problem_anchors or target_anchors or route_anchors):
        return "仅领域相关"
    if not (problem_anchors or target_anchors or route_anchors):
        return "低相关"
    return "需验证"


def score_demand(
    submission: dict,
    demand: dict,
    *,
    tags: dict | None = None,
    capability_profile: dict | None = None,
    profile_anchors: dict[str, set[str]] | None = None,
    recall_score: float = 0.0,
) -> dict:
    tags = normalize_tag_payload(tags or extract_submission_tags_local(submission))
    profile = normalize_technical_profile(capability_profile or build_submission_technical_profile(submission))
    demand_profile = normalize_technical_profile(demand.get("_technical_profile") or build_demand_technical_profile(demand))
    demand_tags = normalize_tag_payload(demand.get("_structured_tags"))
    profile_anchors = profile_anchors or {
        "problem": extract_technical_anchors(
            technical_profile_text(profile, "core_problem", "required_functions", "problem_terms")
        ),
        "target": extract_technical_anchors(
            technical_profile_text(profile, "target", "application_object", "target_terms")
        ),
        "route": extract_technical_anchors(technical_profile_text(profile, "technical_route", "route_terms")),
        "function": extract_functional_anchors(
            technical_profile_text(
                profile,
                "target",
                "core_problem",
                "required_functions",
                "technical_route",
                "application_object",
            )
        ),
    }

    problem_score, problem_overlap = anchored_technical_similarity(
        technical_profile_text(profile, "core_problem", "required_functions", "problem_terms"),
        technical_profile_text(demand_profile, "core_problem", "required_functions", "problem_terms"),
        scale=245,
        left_anchors=profile_anchors.get("problem", set()),
        right_anchors=set(demand.get("_problem_anchors", set())),
    )
    target_score, target_overlap = anchored_technical_similarity(
        technical_profile_text(profile, "target", "application_object", "target_terms"),
        technical_profile_text(demand_profile, "target", "application_object", "target_terms"),
        scale=235,
        left_anchors=profile_anchors.get("target", set()),
        right_anchors=set(demand.get("_target_anchors", set())),
    )
    route_score, route_overlap = anchored_technical_similarity(
        technical_profile_text(profile, "technical_route", "route_terms"),
        technical_profile_text(demand_profile, "technical_route", "route_terms"),
        scale=245,
        left_anchors=profile_anchors.get("route", set()),
        right_anchors=set(demand.get("_route_anchors", set())),
    )
    function_text = technical_profile_text(profile, "required_functions", "problem_terms")
    demand_function_text = technical_profile_text(demand_profile, "required_functions", "problem_terms")
    user_function_anchors = set(profile_anchors.get("function", set()))
    demand_function_anchors = set(demand.get("_function_anchors", set()))
    function_left_anchors = user_function_anchors or extract_technical_anchors(function_text)
    function_right_anchors = demand_function_anchors or extract_technical_anchors(demand_function_text)
    if function_text and demand_function_text:
        function_score, function_overlap = anchored_technical_similarity(
            function_text,
            demand_function_text,
            scale=240,
            left_anchors=function_left_anchors,
            right_anchors=function_right_anchors,
        )
    else:
        function_score, function_overlap = 45, []
    indicator_score, verified_indicators, unverified_indicators = indicator_fit_score(profile, demand_profile)
    maturity_value = maturity_score(clean_text(submission.get("maturity", "")), demand.get(DEMAND_DETAIL_FIELD, ""))
    if not clean_text(submission.get("maturity", "")):
        maturity_value = 48

    tech_overlap = overlap_tags(tags.get("tech_tags", []), demand_tags.get("tech_tags", []))
    scene_overlap = overlap_tags(tags.get("scene_tags", []), demand_tags.get("scene_tags", []))
    industry_overlap = overlap_tags(tags.get("industry_tags", []), demand_tags.get("industry_tags", []))
    total = round(
        problem_score * 0.25
        + target_score * 0.25
        + function_score * 0.20
        + route_score * 0.15
        + indicator_score * 0.10
        + maturity_value * 0.05
    )
    shared_function_anchors = rank_technical_anchors(
        set(function_overlap) | (user_function_anchors & demand_function_anchors),
        6,
    )
    context_adjustment = 0
    if clean_text(submission.get("region", "")) and clean_text(submission.get("region", "")) in demand.get(DEMAND_REGION_FIELD, ""):
        context_adjustment += 2
    if overlap_tags(tags.get("cooperation_tags", []), demand_tags.get("cooperation_tags", [])):
        context_adjustment += 2
    total += min(4, context_adjustment)

    hard_gate = ""
    unique_anchors = set(problem_overlap + target_overlap + function_overlap + route_overlap)
    has_explicit_route = bool(clean_text(submission.get("technical_route", "")))
    has_explicit_problem = bool(clean_text(submission.get("problem", "")))
    has_transfer_evidence = has_explicit_route or bool(profile.get("evidence")) or bool(profile.get("indicators"))
    if not unique_anchors:
        total = min(total, 39)
        hard_gate = "技术标的与核心问题均缺少直接对应证据"
    elif target_overlap and function_score < 20 and problem_score < 30:
        total = min(total, 44)
        hard_gate = "技术对象存在相似性，但需要实现的功能目标不同"
    elif not target_overlap and not has_transfer_evidence:
        total = min(total, 44)
        hard_gate = "需求作用对象不同，且未提供明确技术路线、指标或案例支持技术迁移"
    elif not target_overlap and len(unique_anchors) < 2:
        total = min(total, 44)
        hard_gate = "需求作用对象差异较大，单一共同术语不足以证明技术路线可迁移"
    elif tech_overlap and not (problem_overlap or target_overlap or route_overlap):
        total = min(total, 39)
        hard_gate = "目前仅能确认领域相关，尚不能证明能够解决具体技术任务"
    if route_score < 18 and problem_score < 18 and target_score < 18:
        total = min(total, 35)
        hard_gate = hard_gate or "技术路线与需求任务缺少可验证关联"

    match_type = classify_match_type(
        problem_score,
        target_score,
        route_score,
        tech_overlap,
        problem_overlap,
        target_overlap,
        route_overlap,
        has_explicit_route=has_explicit_route,
        has_explicit_problem=has_explicit_problem,
    )
    confidence = submission_confidence(submission, profile, demand_profile)
    matched_capabilities = display_technical_anchors(
        [*shared_function_anchors, *problem_overlap, *target_overlap, *route_overlap],
        8,
    )
    verified_items = dedupe_keep_order([*matched_capabilities, *verified_indicators], 8)
    unverified_items = dedupe_keep_order(
        [
            *unverified_indicators,
            *([] if profile.get("evidence") else ["缺少样品、检测报告或应用案例证据"]),
            *([] if profile.get("maturity") else ["成果成熟度需要进一步确认"]),
        ],
        6,
    )
    transfer_path = ""
    if match_type == "技术迁移":
        transfer_path = (
            f"可尝试将成果中的{('、'.join(route_overlap[:3]) or '底层技术机理')}迁移到"
            f"{demand_profile.get('target') or demand.get(DEMAND_NAME_FIELD, '')}，需先验证适配改造和关键指标。"
        )
    matched_tags = {
        "tech_tags": tech_overlap,
        "scene_tags": scene_overlap,
        "industry_tags": industry_overlap,
        "cooperation_tags": overlap_tags(tags.get("cooperation_tags", []), demand_tags.get("cooperation_tags", [])),
        "keywords": matched_capabilities,
    }
    reason = build_reason(submission, demand, matched_tags, total, match_type=match_type, hard_gate=hard_gate)
    suggestion = build_suggestion(demand, matched_tags, unverified_items=unverified_items)
    return {
        **sanitize_demand(demand, include_detail=True),
        "demand_detail": public_demand_detail(demand),
        "score": max(0, min(100, total)),
        "confidence": confidence,
        "match_type": match_type,
        "hard_gate": hard_gate,
        "technical_target": clip(demand_profile.get("target") or demand.get(DEMAND_NAME_FIELD, ""), 220),
        "core_problem": clip(demand_profile.get("core_problem", ""), 260),
        "matched_capability": "、".join(matched_capabilities) or "尚未发现足够具体的技术能力对应证据",
        "verified_items": verified_items,
        "unverified_items": unverified_items,
        "hard_conflicts": [],
        "transfer_path": transfer_path,
        "technical_profile": demand_profile,
        "structured_tags": compact_tag_payload(demand_tags),
        "matched_tags": compact_tag_payload(matched_tags),
        "recall_score": round(recall_score, 4),
        "dimensions": {
            "核心问题": problem_score,
            "技术标的": target_score,
            "所需功能": function_score,
            "技术路线": route_score,
            "指标约束": indicator_score,
            "交付成熟度": maturity_value,
        },
        "reason": reason,
        "suggestion": suggestion,
    }


def build_reason(
    submission: dict,
    demand: dict,
    matched_tags: dict,
    score: int,
    *,
    match_type: str = "需验证",
    hard_gate: str = "",
) -> str:
    demand_type = clean_text(demand.get("需求类型", "")) or "技术需求"
    summary_tags = dedupe_keep_order(
        [
            *matched_tags.get("tech_tags", []),
            *matched_tags.get("scene_tags", []),
            *matched_tags.get("industry_tags", []),
            *matched_tags.get("keywords", []),
        ],
        5,
    )
    if summary_tags:
        keyword_text = "、".join(summary_tags[:5])
        lead = f"系统在具体技术任务中发现以下对应点：{keyword_text}。"
    else:
        lead = "当前材料中尚未发现足够具体的技术能力对应证据。"
    if score >= 85:
        level = "匹配度较高"
    elif score >= 70:
        level = "具备进一步沟通价值"
    else:
        level = "可作为备选需求继续核验"
    gate_part = f"但{hard_gate}，因此已限制最高评分。" if hard_gate else ""
    return f"{lead}匹配类型为“{match_type}”，需求类型为{demand_type}。{gate_part}综合判断{level}。"


def build_suggestion(demand: dict, matched_tags: dict, *, unverified_items: list[str] | None = None) -> str:
    keywords = dedupe_keep_order(
        [
            *matched_tags.get("tech_tags", []),
            *matched_tags.get("scene_tags", []),
            *matched_tags.get("industry_tags", []),
            *matched_tags.get("keywords", []),
        ],
        3,
    )
    focus = "、".join(keywords[:3]) if keywords else "样品、指标和应用场景"
    pending = "；优先补充" + "、".join((unverified_items or [])[:3]) if unverified_items else ""
    return f"建议围绕{focus}准备一页技术说明，逐项对应需求标的、技术路线和量化指标{pending}，再由平台人工审核后推进沟通。"


def match_demands(
    store: DemandStore,
    submission: dict,
    limit: int = 12,
    candidate_limit: int = 220,
    tags: dict | None = None,
    capability_profile: dict | None = None,
) -> list[dict]:
    store.refresh_if_changed()
    tags = normalize_tag_payload(tags or extract_submission_tags_local(submission))
    capability_profile = normalize_technical_profile(capability_profile or build_submission_technical_profile(submission))
    profile_tokens = {
        "problem": tokenize(
            technical_profile_text(capability_profile, "core_problem", "problem_terms", "required_functions")
        ),
        "target": tokenize(
            technical_profile_text(capability_profile, "target", "application_object", "target_terms")
        ),
        "route": tokenize(technical_profile_text(capability_profile, "technical_route", "route_terms")),
        "indicator": tokenize(
            technical_profile_text(capability_profile, "indicators", "constraints", "indicator_terms")
        ),
        "problem_anchors": extract_technical_anchors(
            technical_profile_text(capability_profile, "core_problem", "problem_terms", "required_functions")
        ),
        "target_anchors": extract_technical_anchors(
            technical_profile_text(capability_profile, "target", "application_object", "target_terms")
        ),
        "route_anchors": extract_technical_anchors(
            technical_profile_text(capability_profile, "technical_route", "route_terms")
        ),
        "function_anchors": extract_functional_anchors(
            technical_profile_text(
                capability_profile,
                "target",
                "core_problem",
                "required_functions",
                "technical_route",
                "application_object",
            )
        ),
        "all": tokenize(
            build_user_text(
                submission,
                (
                    "title",
                    "application_scene",
                    "summary",
                    "technical_route",
                    "indicators",
                    "evidence",
                    "advantages",
                    "problem",
                    "extra_note",
                    "attachment_note",
                ),
            )
        ),
    }
    coarse_candidates: list[tuple[float, dict]] = []
    for demand in store.demands:
        demand_tags = normalize_tag_payload(demand.get("_structured_tags"))
        lexical = cosine(profile_tokens["all"], demand.get("_tokens", Counter()))
        tag_score = (
            len(overlap_tags(tags.get("tech_tags", []), demand_tags.get("tech_tags", []))) * 0.08
            + len(overlap_tags(tags.get("scene_tags", []), demand_tags.get("scene_tags", []))) * 0.06
            + len(overlap_tags(tags.get("industry_tags", []), demand_tags.get("industry_tags", []))) * 0.035
            + len(overlap_tags(tags.get("keywords", []), demand_tags.get("keywords", []), limit=8)) * 0.025
        )
        function_score = (
            len(
                set(profile_tokens.get("function_anchors", set()))
                & extract_functional_anchors(demand.get("_search_text", ""))
            )
            * 0.16
        )
        coarse_score = lexical + tag_score + function_score
        if coarse_score > 0:
            coarse_candidates.append((coarse_score, demand))

    coarse_candidates.sort(key=lambda item: item[0], reverse=True)
    coarse_limit = min(len(coarse_candidates), max(candidate_limit * 3, 600))
    technical_candidates: list[tuple[float, dict]] = []
    for _, original_demand in coarse_candidates[:coarse_limit]:
        demand = dict(original_demand)
        ensure_demand_technical_profile(demand)
        base = recall_score_demand(submission, demand, tags, capability_profile, profile_tokens)
        if base > 0:
            technical_candidates.append((base, demand))

    technical_candidates.sort(key=lambda item: item[0], reverse=True)
    candidates = technical_candidates[:candidate_limit]
    if not candidates:
        candidates = []
        for original_demand in store.demands[:candidate_limit]:
            demand = dict(original_demand)
            ensure_demand_technical_profile(demand)
            candidates.append((0.0, demand))

    results = [
        score_demand(
            submission,
            demand,
            tags=tags,
            capability_profile=capability_profile,
            profile_anchors={
                "problem": profile_tokens["problem_anchors"],
                "target": profile_tokens["target_anchors"],
                "route": profile_tokens["route_anchors"],
                "function": profile_tokens["function_anchors"],
            },
            recall_score=base,
        )
        for base, demand in candidates
    ]
    results.sort(key=lambda item: item["score"], reverse=True)
    return results[:limit]


def chat_completions_url(base_url: str) -> str:
    base = clean_text(base_url).rstrip("/")
    if base.endswith("/chat/completions"):
        return base
    return f"{base}/chat/completions"


def deepseek_chat(config: dict, messages: list[dict], *, max_tokens: int = 2600) -> str:
    body = {
        "model": config.get("model"),
        "messages": messages,
        "temperature": float(config.get("temperature", 0.2)),
        "max_tokens": max_tokens,
        "stream": False,
        "response_format": {"type": "json_object"},
        "thinking": config.get("thinking") or {"type": "disabled"},
    }
    if config.get("reasoning_effort"):
        body["reasoning_effort"] = config.get("reasoning_effort")

    payload = json.dumps(body, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        chat_completions_url(str(config.get("base_url", ""))),
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {config.get('api_key')}",
        },
    )
    timeout = int(config.get("timeout", 45) or 45)
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")[:500]
        raise RuntimeError(f"DeepSeek API HTTP {exc.code}: {detail}") from exc
    except (urllib.error.URLError, TimeoutError) as exc:
        raise RuntimeError(f"DeepSeek API 请求失败：{exc}") from exc
    choices = data.get("choices") or []
    if not choices:
        raise RuntimeError("DeepSeek API 未返回 choices")
    content = ((choices[0].get("message") or {}).get("content") or "").strip()
    if not content:
        raise RuntimeError("DeepSeek API 返回内容为空")
    return content


def parse_json_object(text: str) -> dict:
    raw = str(text or "").strip()
    if not raw:
        raise ValueError("AI 输出为空")

    candidates: list[str] = [raw]
    fenced = re.sub(r"^\s*```(?:json)?\s*|\s*```\s*$", "", raw, flags=re.IGNORECASE | re.DOTALL).strip()
    if fenced and fenced not in candidates:
        candidates.append(fenced)

    start = fenced.find("{")
    if start >= 0:
        depth = 0
        in_string = False
        escaped = False
        end = -1
        for index, char in enumerate(fenced[start:], start):
            if escaped:
                escaped = False
                continue
            if char == "\\":
                escaped = True
                continue
            if char == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if char == "{":
                depth += 1
            elif char == "}":
                depth -= 1
                if depth == 0:
                    end = index
                    break
        if end > start:
            balanced = fenced[start : end + 1]
            cleaned = re.sub(r",\s*([}\]])", r"\1", balanced)
            cleaned = cleaned.replace("\ufeff", "")
            cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", cleaned)
            if balanced not in candidates:
                candidates.append(balanced)
            if cleaned not in candidates:
                candidates.append(cleaned)

    last_error: Exception | None = None
    for candidate in candidates:
        try:
            data = json.loads(candidate)
            break
        except json.JSONDecodeError as exc:
            last_error = exc
    else:
        raise last_error or ValueError("AI 输出不是合法 JSON")

    if not isinstance(data, dict):
        raise ValueError("AI 输出不是 JSON 对象")
    return data


def build_json_repair_messages(text: str, task_name: str) -> list[dict]:
    system_prompt = (
        "你是 JSON 修复助手。"
        "请把用户提供的文本修复成一个合法 JSON 对象。"
        "不要解释，不要补充无关字段，不要输出 Markdown 代码块。"
    )
    user_prompt = json.dumps(
        {
            "task": task_name,
            "input": str(text or ""),
        },
        ensure_ascii=False,
    )
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def parse_json_object_with_ai(config: dict, text: str, *, task_name: str) -> dict:
    try:
        return parse_json_object(text)
    except Exception as first_error:
        if not ai_is_configured(config):
            raise
        repair_content = deepseek_chat(config, build_json_repair_messages(text, task_name), max_tokens=3200)
        try:
            return parse_json_object(repair_content)
        except Exception as second_error:
            raise RuntimeError(
                f"{task_name} JSON 解析失败：{first_error}; 修复后仍失败：{second_error}"
            ) from second_error


def build_tag_extraction_messages(submission: dict, local_tags: dict) -> list[dict]:
    vocab_payload = {
        "技术标签候选": list(TECH_TAG_VOCAB.keys()),
        "应用标签候选": list(SCENE_TAG_VOCAB.keys()),
        "产业标签候选": list(INDUSTRY_TAG_VOCAB.keys()),
        "合作标签候选": list(COOPERATION_TAG_VOCAB.keys()),
        "成熟度标签候选": list(MATURITY_LABELS.values()),
    }
    system_prompt = (
        "你是技术转移平台的成果能力解析助手。"
        "请把用户成果拆成可用于技术可行性匹配的能力画像，同时输出辅助召回标签。"
        "重点识别成果真正作用的技术标的、能够解决的具体问题、底层原理或工艺路线、"
        "已验证量化指标、样品或案例证据、成熟度和可迁移边界。"
        "行业标签只用于召回，不得替代技术能力分析。只输出合法 JSON。"
    )
    local_profile = build_submission_technical_profile(submission)
    user_payload = {
        "submission": compact_submission(submission, full_text_chars=12000),
        "local_seed_tags": compact_tag_payload(local_tags),
        "local_seed_capability_profile": local_profile,
        "standard_vocab": vocab_payload,
        "required_json_schema": {
            "技术标签": ["最多6个"],
            "应用标签": ["最多6个"],
            "产业标签": ["最多6个"],
            "合作标签": ["最多4个"],
            "成熟度标签": "从候选中选1个，无法判断可留空",
            "关键词": ["最多8个"],
            "成果能力画像": {
                "技术标的": "成果实际作用或交付的对象",
                "核心问题": "成果已经能够解决的具体技术问题",
                "所需功能": ["成果能提供的关键功能，最多6项"],
                "技术路线": "底层机理、材料、工艺、算法或装备路线",
                "已验证指标": ["只填写材料中明确给出的量化指标，最多8项"],
                "限制条件": ["适用边界或限制，最多6项"],
                "应用对象": "适用产品、设备、材料或环境",
                "可交付物": ["样品、配方、设备、系统、工艺包等，最多6项"],
                "证据": ["样品、检测、专利、客户案例等明确证据，最多6项"],
                "成熟度": "概念、实验室、小试、中试或量产",
                "技术标的词": ["最多12个具体词"],
                "问题词": ["最多12个具体词"],
                "路线词": ["最多12个具体词"],
                "指标词": ["最多12个具体词"],
            },
        },
    }
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
    ]


def normalize_ai_tag_payload(payload: dict) -> dict:
    return normalize_tag_payload(
        {
            "tech_tags": normalize_vocab_values(payload.get("技术标签"), TECH_ALIAS_LOOKUP, TECH_TAG_VOCAB, limit=6),
            "scene_tags": normalize_vocab_values(payload.get("应用标签"), SCENE_ALIAS_LOOKUP, SCENE_TAG_VOCAB, limit=6),
            "industry_tags": normalize_vocab_values(payload.get("产业标签"), INDUSTRY_ALIAS_LOOKUP, INDUSTRY_TAG_VOCAB, limit=6),
            "cooperation_tags": normalize_vocab_values(payload.get("合作标签"), COOP_ALIAS_LOOKUP, COOPERATION_TAG_VOCAB, limit=4),
            "keywords": dedupe_keep_order(list_values(payload.get("关键词")), 8),
            "region_tokens": [],
            "maturity_label": clean_text(payload.get("成熟度标签", "")),
        }
    )


def normalize_ai_technical_profile(payload: dict) -> dict:
    profile = payload.get("成果能力画像") if isinstance(payload.get("成果能力画像"), dict) else {}
    return normalize_technical_profile(
        {
            "target": profile.get("技术标的"),
            "core_problem": profile.get("核心问题"),
            "required_functions": profile.get("所需功能"),
            "technical_route": profile.get("技术路线"),
            "indicators": profile.get("已验证指标"),
            "constraints": profile.get("限制条件"),
            "application_object": profile.get("应用对象"),
            "deliverables": profile.get("可交付物"),
            "evidence": profile.get("证据"),
            "maturity": profile.get("成熟度"),
            "target_terms": profile.get("技术标的词"),
            "problem_terms": profile.get("问题词"),
            "route_terms": profile.get("路线词"),
            "indicator_terms": profile.get("指标词"),
        }
    )


def extract_tags_with_ai(config: dict, submission: dict, local_tags: dict) -> tuple[dict, dict]:
    local_tags = normalize_tag_payload(local_tags)
    local_profile = build_submission_technical_profile(submission)
    if not ai_is_configured(config):
        return local_tags, {
            "used_ai": False,
            "source": "local",
            "local_tags": compact_tag_payload(local_tags),
            "merged_tags": compact_tag_payload(local_tags),
            "capability_profile": local_profile,
            "message": "未配置 DeepSeek API，已使用本地标签抽取。",
        }
    try:
        content = deepseek_chat(config, build_tag_extraction_messages(submission, local_tags), max_tokens=3000)
        ai_payload = parse_json_object_with_ai(config, content, task_name="标签抽取")
        ai_tags = normalize_ai_tag_payload(ai_payload)
        ai_profile = normalize_ai_technical_profile(ai_payload)
        merged_tags = merge_tag_profiles(local_tags, ai_tags)
        merged_profile = merge_technical_profiles(local_profile, ai_profile)
        return merged_tags, {
            "used_ai": True,
            "source": "local+ai",
            "local_tags": compact_tag_payload(local_tags),
            "ai_tags": compact_tag_payload(ai_tags),
            "merged_tags": compact_tag_payload(merged_tags),
            "capability_profile": merged_profile,
            "message": "已使用 DeepSeek 完成成果能力画像与辅助标签抽取。",
        }
    except Exception as exc:
        return local_tags, {
            "used_ai": False,
            "source": "local",
            "local_tags": compact_tag_payload(local_tags),
            "merged_tags": compact_tag_payload(local_tags),
            "capability_profile": local_profile,
            "message": f"AI 标签抽取失败，已退回本地标签：{exc}",
        }


def compact_submission(submission: dict, *, full_text_chars: int = 3500) -> dict:
    return {
        "成果名称": clean_text(submission.get("title") or submission.get("achievement_name")),
        "成果全文": clip(submission.get("achievement_text") or submission.get("full_text", ""), full_text_chars),
        "技术领域": clean_text(submission.get("tech_field")),
        "应用场景": clip(submission.get("application_scene", ""), 180),
        "技术成果摘要": clip(submission.get("summary", ""), 300),
        "技术路线": clip(submission.get("technical_route", ""), 260),
        "量化指标": clip(submission.get("indicators", ""), 220),
        "验证证据": clip(submission.get("evidence", ""), 220),
        "核心优势": clip(submission.get("advantages", ""), 180),
        "解决问题": clip(submission.get("problem", ""), 220),
        "成熟度": clean_text(submission.get("maturity")),
        "合作方式": clean_text(submission.get("cooperation")),
        "所在地区": clean_text(submission.get("region")),
        "补充说明或相关链接": clip(submission.get("extra_note") or submission.get("attachment_note", ""), 160),
    }


def capability_profile_form_fields(profile: dict | None) -> dict:
    """Expose AI-parsed profile values for a user-facing preview or correction form."""
    normalized = normalize_technical_profile(profile)
    return {
        "application_scene": normalized.get("application_object", ""),
        "summary": "；".join(
            value
            for value in [normalized.get("target", ""), normalized.get("core_problem", "")]
            if value
        ),
        "technical_route": normalized.get("technical_route", ""),
        "indicators": "；".join(normalized.get("indicators") or []),
        "evidence": "；".join(normalized.get("evidence") or []),
        "problem": normalized.get("core_problem", ""),
        "maturity": normalized.get("maturity", ""),
    }


def compact_ai_technical_profile(profile: dict | None) -> dict:
    normalized = normalize_technical_profile(profile)
    return {
        "target": clip(normalized.get("target", ""), 160),
        "core_problem": clip(normalized.get("core_problem", ""), 280),
        "required_functions": (normalized.get("required_functions") or [])[:4],
        "technical_route": clip(normalized.get("technical_route", ""), 300),
        "indicators": (normalized.get("indicators") or [])[:5],
        "constraints": [clip(item, 140) for item in (normalized.get("constraints") or [])[:4]],
        "application_object": clip(normalized.get("application_object", ""), 160),
        "deliverables": (normalized.get("deliverables") or [])[:4],
        "evidence": [clip(item, 140) for item in (normalized.get("evidence") or [])[:3]],
        "maturity": clean_text(normalized.get("maturity", "")),
    }


def compact_candidate(result: dict) -> dict:
    return {
        "需求ID": result.get("demand_id", ""),
        "需求名称": result.get("name", ""),
        "技术领域": clip(result.get("tech_field", ""), 160),
        "需求类型": result.get("demand_type", ""),
        "合作方式": result.get("cooperation_mode", ""),
        "需求技术任务书": compact_ai_technical_profile(result.get("technical_profile") or {}),
        "关键需求原文": clip(result.get("demand_detail") or result.get("detail_summary", ""), 400),
        "本地匹配分": result.get("score", 0),
        "本地判断可信度": result.get("confidence", 0),
        "本地匹配类型": result.get("match_type", ""),
        "本地硬门槛": result.get("hard_gate", ""),
        "本地维度评分": result.get("dimensions", {}),
    }


def build_ai_messages(
    submission: dict,
    local_results: list[dict],
    tag_profile: dict | None = None,
    capability_profile: dict | None = None,
) -> list[dict]:
    system_prompt = """
你是技术转移平台的高级技术经理人与可行性评审专家。请在一次处理中完成两项任务：
1. 根据用户原始材料校正本地生成的成果能力画像。
2. 判断成果能力能否解决每条候选需求正文中的具体技术任务并完成精排。

请逐条比较：
1. 核心问题匹配（25%）：成果是否解决需求真正的技术瓶颈。
2. 技术标的匹配（25%）：成果作用对象、产品或交付对象是否与需求一致。
3. 所需功能匹配（20%）：成果已经具备的功能是否对应需求希望实现的功能目标。
4. 技术路线可行性（15%）：底层机理、材料、工艺、算法或装备路线是否适用。
5. 指标与约束满足度（10%）：需求量化指标和限制条件是否已有证据满足。
6. 成熟度与交付能力（5%）：样品、中试、工程化和交付能力是否符合需求阶段。

规则：
- 只能使用候选需求中的信息，不要编造不存在的需求。
- 行业、地区、宽泛领域标签只能辅助理解，不能作为高分依据。
- 同一种材料、设备或算法如果解决的功能目标不同（例如导热与导电），不能判为“直接解决”；除非技术路线和验证指标能够证明可满足需求，否则总分不宜超过65。
- “稳定性、性能、效率、加工、管理、控制”等通用表述不能单独作为技术迁移依据。
- 如果技术标的和核心问题均不匹配，总分不得超过40。
- 如果仅行业相同、没有具体技术证据，总分不得超过45，匹配类型必须为“仅领域相关”。
- 如果存在明确技术路线冲突或强制指标无法满足，总分不得超过35，并写入硬性冲突。
- 允许跨行业技术迁移，但必须说明底层机理、迁移路径、适配改造和需要验证的条件。
- 材料未提供某项指标时，应标记“未验证”，不得自行推断已经满足。
- 技术匹配度与判断可信度必须分开。资料不完整可以降低可信度，但不能伪造证据。
- 不要输出需求方联系方式、手机号、联系人或外部详情页链接。
- 总分、可信度和六个维度分数均为 0 到 100 的整数。
- reason 要像技术经理人写给用户看的中文说明，具体但简洁。
- suggestion 要给出下一步撮合建议。
- 只评估给出的候选需求，不扩写背景知识。每项说明尽量控制在80字以内。
- 必须输出合法 json，格式如下：
{
  "capability_profile": {
    "target": "成果实际作用或交付的对象",
    "core_problem": "成果已经能够解决的具体技术问题",
    "required_functions": ["最多4项"],
    "technical_route": "底层机理、材料、工艺、算法或装备路线",
    "indicators": ["材料中明确提供的量化指标，最多5项"],
    "constraints": ["适用边界或限制，最多4项"],
    "application_object": "适用产品、设备、材料或环境",
    "deliverables": ["样品、配方、设备、系统、工艺包等，最多4项"],
    "evidence": ["样品、检测或案例等明确证据，最多3项"],
    "maturity": "概念、实验室、小试、中试或量产"
  },
  "results": [
    {
      "demand_id": "候选需求ID",
      "score": 88,
      "confidence": 72,
      "match_type": "直接解决/关键组件/技术迁移/仅领域相关/低相关/需验证",
      "dimensions": {"核心问题": 90, "技术标的": 86, "所需功能": 88, "技术路线": 82, "指标约束": 70, "交付成熟度": 82},
      "technical_target": "需求要研发或交付的具体对象",
      "core_problem": "需求真正要解决的技术问题",
      "matched_capability": "成果能够提供的对应技术能力",
      "verified_items": ["已有证据支持的对应项"],
      "unverified_items": ["尚需验证的指标或条件"],
      "hard_conflicts": ["明确冲突，没有则为空数组"],
      "transfer_path": "跨行业迁移时填写，否则为空",
      "reason": "中文匹配理由",
      "suggestion": "中文合作建议"
    }
  ]
}
"""
    payload = {
        "submission": compact_submission(submission),
        "local_capability_profile": compact_ai_technical_profile(
            capability_profile or build_submission_technical_profile(submission)
        ),
        "structured_tags": compact_tag_payload(tag_profile),
        "candidates": [compact_candidate(item) for item in local_results],
    }
    user_prompt = "请校正成果能力画像并对以下候选需求做技术精排，只输出 json：\n" + json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def to_int_score(value: object, default: int = 0) -> int:
    try:
        number = int(round(float(value)))
    except (TypeError, ValueError):
        number = default
    return max(0, min(100, number))


def to_float_score(value: object, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def merge_ai_results(local_results: list[dict], ai_payload: dict, config: dict) -> list[dict]:
    by_id = {item.get("demand_id"): item for item in local_results}
    merged: list[dict] = []
    seen: set[str] = set()
    for ai_item in ai_payload.get("results", []):
        if not isinstance(ai_item, dict):
            continue
        demand_id = clean_text(ai_item.get("demand_id"))
        if demand_id not in by_id or demand_id in seen:
            continue
        base = dict(by_id[demand_id])
        local_score = to_int_score(base.get("score"), 0)
        local_confidence = to_int_score(base.get("confidence"), 0)
        local_recall_score = to_float_score(base.get("recall_score"), 0.0)
        local_dimensions = base.get("dimensions") if isinstance(base.get("dimensions"), dict) else {}
        local_hard_gate = clean_text(base.get("hard_gate", ""))
        dimensions = ai_item.get("dimensions") if isinstance(ai_item.get("dimensions"), dict) else {}
        score = to_int_score(ai_item.get("score"), base.get("score", 0))
        confidence = to_int_score(ai_item.get("confidence"), base.get("confidence", 50))
        match_type = clean_text(ai_item.get("match_type")) or base.get("match_type", "需验证")
        hard_conflicts = dedupe_keep_order(list_values(ai_item.get("hard_conflicts")), 6)
        base["dimensions"] = {
            "核心问题": to_int_score(dimensions.get("核心问题"), base.get("dimensions", {}).get("核心问题", 0)),
            "技术标的": to_int_score(dimensions.get("技术标的"), base.get("dimensions", {}).get("技术标的", 0)),
            "所需功能": to_int_score(dimensions.get("所需功能"), base.get("dimensions", {}).get("所需功能", 0)),
            "技术路线": to_int_score(dimensions.get("技术路线"), base.get("dimensions", {}).get("技术路线", 0)),
            "指标约束": to_int_score(dimensions.get("指标约束"), base.get("dimensions", {}).get("指标约束", 0)),
            "交付成熟度": to_int_score(
                dimensions.get("交付成熟度"), base.get("dimensions", {}).get("交付成熟度", 0)
            ),
        }
        if base["dimensions"]["技术标的"] < 25 and base["dimensions"]["核心问题"] < 25:
            score = min(score, 40)
            base["hard_gate"] = "技术标的与核心问题均缺少直接对应证据"
        elif (
            base["dimensions"]["技术标的"] < 20
            and base["dimensions"]["核心问题"] < 65
            and base["dimensions"]["技术路线"] < 50
        ):
            score = min(score, 44)
            base["hard_gate"] = "需求作用对象差异较大，现有技术路线证据不足以支持迁移"
        if hard_conflicts:
            score = min(score, 35)
            base["hard_gate"] = "存在明确技术路线或强制指标冲突"
        if match_type in {"仅领域相关", "低相关"}:
            score = min(score, 45 if match_type == "仅领域相关" else 40)
        weak_local_evidence = (
            local_recall_score < 0.03
            or local_score < 45
            or "缺少直接对应证据" in local_hard_gate
            or "不足以支持迁移" in local_hard_gate
        )
        local_core = to_int_score(local_dimensions.get("核心问题"), 0)
        local_target = to_int_score(local_dimensions.get("技术标的"), 0)
        local_function = to_int_score(local_dimensions.get("所需功能"), 0)
        local_route = to_int_score(local_dimensions.get("技术路线"), 0)
        weak_technical_dimensions = (
            local_core < 35 and local_target < 35 and local_function < 35 and local_route < 45
        )
        if weak_local_evidence and weak_technical_dimensions:
            score = min(score, 44)
            confidence = min(confidence, max(local_confidence, 55))
            match_type = "需验证" if match_type not in {"低相关", "仅领域相关"} else match_type
            base["hard_gate"] = local_hard_gate or "本地技术证据不足，AI 复核不得上调为高匹配"
        elif local_score < 60 and score - local_score > 18:
            score = min(score, local_score + 18)
        base["score"] = score
        base["confidence"] = confidence
        base["match_type"] = match_type
        base["technical_target"] = clip(ai_item.get("technical_target") or base.get("technical_target", ""), 260)
        base["core_problem"] = clip(ai_item.get("core_problem") or base.get("core_problem", ""), 260)
        base["matched_capability"] = clip(
            ai_item.get("matched_capability") or base.get("matched_capability", ""), 300
        )
        base["verified_items"] = dedupe_keep_order(
            [*list_values(ai_item.get("verified_items")), *list_values(base.get("verified_items"))], 8
        )
        base["unverified_items"] = dedupe_keep_order(
            [*list_values(ai_item.get("unverified_items")), *list_values(base.get("unverified_items"))], 8
        )
        base["hard_conflicts"] = hard_conflicts
        base["transfer_path"] = clip(ai_item.get("transfer_path") or base.get("transfer_path", ""), 300)
        base["reason"] = clip(ai_item.get("reason") or base.get("reason", ""), 380)
        base["suggestion"] = clip(ai_item.get("suggestion") or base.get("suggestion", ""), 260)
        base["scoring_source"] = "成果能力画像 + AI技术精排"
        base["ai_model"] = clean_text(config.get("model", ""))
        merged.append(base)
        seen.add(demand_id)

    for item in local_results:
        if item.get("demand_id") not in seen:
            fallback = dict(item)
            fallback["scoring_source"] = "本地技术任务评分"
            merged.append(fallback)
    merged.sort(key=lambda item: (item.get("score", 0), item.get("confidence", 0)), reverse=True)
    return merged


def refine_matches_with_ai(
    config: dict,
    submission: dict,
    local_results: list[dict],
    *,
    tag_profile: dict | None = None,
    tag_meta: dict | None = None,
    capability_profile: dict | None = None,
) -> tuple[list[dict], dict]:
    if not ai_is_configured(config):
        for item in local_results:
            item["scoring_source"] = "本地技术任务评分"
        return local_results, {
            "used_ai": False,
            "match_mode": "ai",
            "message": "未配置 DeepSeek API，已使用本地技术任务解析与五维评分。",
            "capability_profile": normalize_technical_profile(capability_profile),
        }

    try:
        ai_candidates = local_results[:6]
        messages = build_ai_messages(submission, ai_candidates, tag_profile, capability_profile)
        content = deepseek_chat(config, messages, max_tokens=2800)
        # Keep the optimized path to exactly one external AI call. If the model
        # returns malformed JSON, fall back to local scoring instead of making
        # a second repair request.
        ai_payload = parse_json_object(content)
        ai_profile = ai_payload.get("capability_profile") if isinstance(ai_payload.get("capability_profile"), dict) else {}
        corrected_profile = merge_technical_profiles(capability_profile or {}, ai_profile)
        refined = merge_ai_results(local_results, ai_payload, config)
        ai_ranked_count = sum(
            1 for item in refined if item.get("scoring_source") == "成果能力画像 + AI技术精排"
        )
        if ai_ranked_count == 0:
            for item in local_results:
                item["scoring_source"] = "本地技术任务评分"
            return local_results, {
                "used_ai": False,
                "match_mode": "ai",
                "message": "AI 深度复核未返回有效候选，当前展示本地技术评分结果。",
                "model": clean_text(config.get("model", "")),
                "tag_extraction": tag_meta or {},
                "structured_tags": compact_tag_payload(tag_profile),
                "capability_profile": corrected_profile,
                "ai_candidate_count": len(ai_candidates),
            }
        return refined, {
            "used_ai": True,
            "match_mode": "ai",
            "message": "已基于成果材料生成能力画像，并对前 6 条候选需求进行 AI 技术复核。",
            "model": clean_text(config.get("model", "")),
            "tag_extraction": tag_meta or {},
            "structured_tags": compact_tag_payload(tag_profile),
            "capability_profile": corrected_profile,
            "ai_candidate_count": len(ai_candidates),
        }
    except Exception as exc:
        print(f"DeepSeek AI review failed: {exc}")
        for item in local_results:
            item["scoring_source"] = "本地技术任务评分"
        lowered_error = str(exc).lower()
        if "timed out" in lowered_error or "timeout" in lowered_error:
            message = "AI 深度复核暂未在时限内完成，当前展示本地技术评分结果，可稍后重新复核。"
        else:
            message = "AI 深度复核暂时不可用，当前展示本地技术评分结果。"
        return local_results, {
            "used_ai": False,
            "match_mode": "ai",
            "message": message,
            "capability_profile": normalize_technical_profile(capability_profile),
            "ai_candidate_count": min(6, len(local_results)),
        }


def use_quick_match(local_results: list[dict]) -> tuple[list[dict], dict]:
    for item in local_results:
        item["scoring_source"] = "本地技术任务评分"
    return local_results, {
        "used_ai": False,
        "match_mode": "quick",
        "message": "已使用本地技术任务解析完成快速匹配，重点比较技术标的、核心问题、技术路线和指标约束，未调用 DeepSeek API。",
    }


def ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def append_jsonl(name: str, payload: dict) -> None:
    ensure_data_dir()
    path = DATA_DIR / name
    with path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(payload, ensure_ascii=False) + "\n")


def read_jsonl(name: str, limit: int = 100) -> list[dict]:
    path = DATA_DIR / name
    if not path.exists():
        return []
    rows: list[dict] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    rows.reverse()
    return rows[:limit]


def json_dumps(payload: object) -> str:
    return json.dumps(payload, ensure_ascii=False)


def using_postgres() -> bool:
    return DATABASE_URL.lower().startswith(("postgres://", "postgresql://"))


def db_prepare(sql: str) -> str:
    if using_postgres():
        return sql.replace("?", "%s")
    return sql


def db_execute(conn: object, sql: str, params: tuple | list = ()) -> object:
    return conn.execute(db_prepare(sql), tuple(params))


def first_value(row: object) -> object:
    if row is None:
        return 0
    if isinstance(row, dict):
        return next(iter(row.values()), 0)
    return row[0]


def db_connect() -> object:
    ensure_data_dir()
    if using_postgres():
        try:
            import psycopg
            from psycopg.rows import dict_row
        except ImportError as exc:
            raise RuntimeError("线上数据库需要安装 psycopg：请先执行 pip install -r requirements.txt") from exc
        return psycopg.connect(DATABASE_URL, row_factory=dict_row)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def demand_public_payload(row: dict) -> dict:
    return {field: clean_text(row.get(field)) for field in DEMAND_FIELDS}


def database_demands_version() -> str:
    try:
        with db_connect() as conn:
            row = db_execute(
                conn,
                """
                SELECT
                    (SELECT COUNT(*) FROM demands) AS count,
                    (SELECT MAX(updated_at) FROM demands) AS updated_at,
                    (SELECT COUNT(*) FROM demand_analyses) AS analysis_count,
                    (SELECT MAX(updated_at) FROM demand_analyses) AS analysis_updated_at
                """,
            ).fetchone()
    except Exception:
        return ""
    if not row:
        return ""
    item = dict(row)
    count = int(item.get("count") or 0)
    if count <= 0:
        return ""
    return (
        f"db:{count}:{clean_text(item.get('updated_at'))}:"
        f"{int(item.get('analysis_count') or 0)}:{clean_text(item.get('analysis_updated_at'))}"
    )


def load_demands_from_database() -> list[dict]:
    try:
        with db_connect() as conn:
            rows = db_execute(
                conn,
                """
                SELECT demand_json
                FROM demands
                ORDER BY updated_at DESC, first_seen_at DESC
                """,
            ).fetchall()
    except Exception:
        return []

    analysis_map = load_demand_analysis_map()
    demands: list[dict] = []
    for row in rows:
        item = dict(row)
        try:
            payload = json.loads(clean_text(item.get("demand_json")))
        except json.JSONDecodeError:
            continue
        cleaned = prepare_demand_row(payload if isinstance(payload, dict) else {})
        if cleaned:
            demand_id = clean_text(cleaned.get(DEMAND_ID_FIELD))
            analysis = analysis_map.get(demand_id) or {}
            if analysis and clean_text(analysis.get("content_hash")) == demand_content_hash(cleaned):
                profile = normalize_technical_profile(analysis.get("profile"))
                if any(profile.values()):
                    cleaned["_technical_profile"] = profile
                    cleaned["_analysis_source"] = clean_text(analysis.get("source"))
                    cleaned["_analysis_version"] = clean_text(analysis.get("analysis_version"))
                    cleaned["_analysis_quality"] = int(analysis.get("quality_score") or 0)
                    profile_search_text = technical_profile_text(
                        profile,
                        "target",
                        "core_problem",
                        "required_functions",
                        "technical_route",
                        "indicators",
                        "constraints",
                        "application_object",
                        "deliverables",
                        "target_terms",
                        "problem_terms",
                        "route_terms",
                        "indicator_terms",
                    )
                    cleaned["_search_text"] = " ".join(
                        [cleaned.get("_search_text", ""), profile_search_text]
                    ).strip()
                    cleaned["_tokens"] = tokenize(cleaned["_search_text"])
            demands.append(cleaned)
    return demands


def existing_demand_ids() -> set[str]:
    try:
        with db_connect() as conn:
            rows = db_execute(conn, "SELECT demand_id FROM demands").fetchall()
            return {clean_text(dict(row).get("demand_id")) for row in rows if clean_text(dict(row).get("demand_id"))}
    except Exception:
        return set()


def load_demand_analysis_map() -> dict[str, dict]:
    try:
        with db_connect() as conn:
            rows = db_execute(
                conn,
                """
                SELECT demand_id, content_hash, analysis_version, status, source, model,
                       profile_json, local_profile_json, quality_score, error, analyzed_at, updated_at
                FROM demand_analyses
                """,
            ).fetchall()
    except Exception:
        return {}

    analyses: dict[str, dict] = {}
    for row in rows:
        item = dict(row)
        demand_id = clean_text(item.get("demand_id"))
        if not demand_id:
            continue
        try:
            profile = json.loads(clean_text(item.get("profile_json")) or "{}")
        except json.JSONDecodeError:
            profile = {}
        try:
            local_profile = json.loads(clean_text(item.get("local_profile_json")) or "{}")
        except json.JSONDecodeError:
            local_profile = {}
        analyses[demand_id] = {
            **item,
            "profile": profile if isinstance(profile, dict) else {},
            "local_profile": local_profile if isinstance(local_profile, dict) else {},
        }
    return analyses


def upsert_demand_analysis(record: dict) -> None:
    timestamp = clean_text(record.get("updated_at")) or now_iso()
    analyzed_at = clean_text(record.get("analyzed_at")) or timestamp
    profile = normalize_technical_profile(record.get("profile") or {})
    local_profile = normalize_technical_profile(record.get("local_profile") or profile)
    with db_connect() as conn:
        db_execute(
            conn,
            """
            INSERT INTO demand_analyses (
                demand_id, content_hash, analysis_version, status, source, model,
                profile_json, local_profile_json, quality_score, error, analyzed_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (demand_id) DO UPDATE SET
                content_hash = excluded.content_hash,
                analysis_version = excluded.analysis_version,
                status = excluded.status,
                source = excluded.source,
                model = excluded.model,
                profile_json = excluded.profile_json,
                local_profile_json = excluded.local_profile_json,
                quality_score = excluded.quality_score,
                error = excluded.error,
                analyzed_at = excluded.analyzed_at,
                updated_at = excluded.updated_at
            """,
            (
                clean_text(record.get("demand_id")),
                clean_text(record.get("content_hash")),
                clean_text(record.get("analysis_version")) or DEMAND_ANALYSIS_VERSION,
                clean_text(record.get("status")) or "ready",
                clean_text(record.get("source")) or "local",
                clean_text(record.get("model")),
                json_dumps(profile),
                json_dumps(local_profile),
                max(0, min(100, int(record.get("quality_score") or 0))),
                clip(record.get("error", ""), 1000),
                analyzed_at,
                timestamp,
            ),
        )


def demand_analysis_stats() -> dict:
    try:
        with db_connect() as conn:
            rows = db_execute(
                conn,
                """
                SELECT status, source, COUNT(*) AS count
                FROM demand_analyses
                GROUP BY status, source
                """,
            ).fetchall()
    except Exception:
        return {"total": 0, "groups": []}
    groups = [dict(row) for row in rows]
    return {
        "total": sum(int(item.get("count") or 0) for item in groups),
        "groups": groups,
    }


def save_demand_rows_to_database(rows: list[dict], *, update_existing: bool = True) -> dict:
    if not rows:
        return {"inserted": 0, "updated": 0, "skipped": 0}

    timestamp = now_iso()
    inserted = 0
    updated = 0
    skipped = 0
    with db_connect() as conn:
        for raw in rows:
            row = demand_public_payload(raw)
            demand_id = clean_text(row.get("需求ID"))
            name = clean_text(row.get("需求名称"))
            if not demand_id or not name:
                skipped += 1
                continue
            if not update_existing:
                cursor = db_execute(
                    conn,
                    """
                    INSERT INTO demands (
                        demand_id, demand_no, name, demand_json, source, source_url, first_seen_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (demand_id) DO NOTHING
                    """,
                    (
                        demand_id,
                        row.get("需求编号", ""),
                        name,
                        json_dumps(row),
                        "jstec",
                        row.get("详情页链接", ""),
                        timestamp,
                        timestamp,
                    ),
                )
                if getattr(cursor, "rowcount", 0):
                    inserted += 1
                else:
                    skipped += 1
                continue

            existing = db_execute(conn, "SELECT 1 FROM demands WHERE demand_id = ?", (demand_id,)).fetchone()
            db_execute(
                conn,
                """
                INSERT INTO demands (
                    demand_id, demand_no, name, demand_json, source, source_url, first_seen_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (demand_id) DO UPDATE SET
                    demand_no = excluded.demand_no,
                    name = excluded.name,
                    demand_json = excluded.demand_json,
                    source = excluded.source,
                    source_url = excluded.source_url,
                    updated_at = excluded.updated_at
                """,
                (
                    demand_id,
                    row.get("需求编号", ""),
                    name,
                    json_dumps(row),
                    "jstec",
                    row.get("详情页链接", ""),
                    timestamp,
                    timestamp,
                ),
            )
            if existing:
                updated += 1
            else:
                inserted += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def migrate_demands_file_to_database() -> None:
    if not DEMANDS_FILE.exists():
        return
    try:
        if existing_demand_ids():
            return
        rows = [demand_public_payload(row) for row in load_demands_from_file(DEMANDS_FILE)]
        save_demand_rows_to_database(rows, update_existing=False)
    except Exception as exc:
        print(f"需求库文件迁移到数据库失败，仍将使用本地文件：{exc}")


def init_database() -> None:
    ensure_data_dir()
    schema_sql = """
    CREATE TABLE IF NOT EXISTS submissions (
        submission_id TEXT PRIMARY KEY,
        created_at TEXT NOT NULL,
        submission_json TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS matches (
        match_id TEXT PRIMARY KEY,
        submission_id TEXT,
        created_at TEXT NOT NULL,
        ai_meta_json TEXT NOT NULL,
        results_json TEXT NOT NULL,
        submission_json TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS intents (
        intent_id TEXT PRIMARY KEY,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        status TEXT NOT NULL,
        agreement_version TEXT,
        submission_id TEXT,
        contact_json TEXT NOT NULL,
        message TEXT,
        attachment_note TEXT,
        selected_result_json TEXT NOT NULL,
        followup_note TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS intent_status_logs (
        log_id TEXT PRIMARY KEY,
        intent_id TEXT NOT NULL,
        old_status TEXT,
        new_status TEXT NOT NULL,
        note TEXT,
        operator TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS match_followups (
        followup_id TEXT PRIMARY KEY,
        match_id TEXT NOT NULL,
        submission_id TEXT,
        demand_id TEXT NOT NULL,
        status TEXT NOT NULL,
        contact_note TEXT DEFAULT '',
        project_progress TEXT DEFAULT '',
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        UNIQUE (match_id, demand_id)
    );

    CREATE TABLE IF NOT EXISTS demands (
        demand_id TEXT PRIMARY KEY,
        demand_no TEXT,
        name TEXT NOT NULL,
        demand_json TEXT NOT NULL,
        source TEXT DEFAULT 'jstec',
        source_url TEXT,
        first_seen_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS demand_analyses (
        demand_id TEXT PRIMARY KEY,
        content_hash TEXT NOT NULL,
        analysis_version TEXT NOT NULL,
        status TEXT NOT NULL,
        source TEXT NOT NULL,
        model TEXT DEFAULT '',
        profile_json TEXT NOT NULL,
        local_profile_json TEXT NOT NULL,
        quality_score INTEGER DEFAULT 0,
        error TEXT DEFAULT '',
        analyzed_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );

    CREATE INDEX IF NOT EXISTS idx_demands_updated_at ON demands (updated_at);
    CREATE INDEX IF NOT EXISTS idx_demand_analyses_status ON demand_analyses (status);
    CREATE INDEX IF NOT EXISTS idx_demand_analyses_version ON demand_analyses (analysis_version);
    CREATE INDEX IF NOT EXISTS idx_match_followups_match_id ON match_followups (match_id);
    """
    with db_connect() as conn:
        if using_postgres():
            for statement in schema_sql.split(";"):
                statement = statement.strip()
                if statement:
                    db_execute(conn, statement)
        else:
            conn.executescript(schema_sql)
    migrate_database_schema()
    migrate_jsonl_to_database()
    migrate_demands_file_to_database()


def column_exists(conn: object, table: str, column: str) -> bool:
    if using_postgres():
        row = db_execute(
            conn,
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = ? AND column_name = ?
            LIMIT 1
            """,
            (table, column),
        ).fetchone()
        return row is not None
    rows = db_execute(conn, f"PRAGMA table_info({table})").fetchall()
    return any(dict(row).get("name") == column for row in rows)


def add_column_if_missing(conn: object, table: str, column: str, definition: str) -> None:
    if column_exists(conn, table, column):
        return
    db_execute(conn, f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def migrate_database_schema() -> None:
    with db_connect() as conn:
        add_column_if_missing(conn, "intents", "query_code", "TEXT")
        add_column_if_missing(conn, "intents", "public_note", "TEXT DEFAULT ''")
        add_column_if_missing(conn, "intents", "progress_json", "TEXT DEFAULT ''")
        add_column_if_missing(conn, "intents", "deal_amount", "TEXT DEFAULT ''")
        add_column_if_missing(conn, "intents", "deal_note", "TEXT DEFAULT ''")

        rows = db_execute(
            conn,
            """
            SELECT intent_id, created_at, query_code, progress_json, public_note
            FROM intents
            """,
        ).fetchall()
        for row in rows:
            item = dict(row)
            updates: list[str] = []
            params: list[str] = []
            query_code = clean_text(item.get("query_code"))
            if not query_code:
                query_code = generate_unique_query_code(conn)
                updates.append("query_code = ?")
                params.append(query_code)
            if not clean_text(item.get("progress_json")):
                updates.append("progress_json = ?")
                params.append(json_dumps(default_progress(item.get("created_at"))))
            if not clean_text(item.get("public_note")):
                updates.append("public_note = ?")
                params.append(f"已收到合作意向，{PUBLIC_RESPONSE_PROMISE}")
            if updates:
                params.append(clean_text(item.get("intent_id")))
                db_execute(conn, f"UPDATE intents SET {', '.join(updates)} WHERE intent_id = ?", params)


def generate_query_code() -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "TN-" + "".join(secrets.choice(alphabet) for _ in range(8))


def generate_unique_query_code(conn: object | None = None) -> str:
    own_conn = conn is None
    if own_conn:
        conn = db_connect()
    try:
        for _ in range(30):
            code = generate_query_code()
            row = db_execute(conn, "SELECT 1 FROM intents WHERE query_code = ?", (code,)).fetchone()
            if row is None:
                return code
    finally:
        if own_conn and conn is not None:
            conn.close()
    return "TN-" + uuid.uuid4().hex[:8].upper()


def default_progress(created_at: object = "") -> list[dict]:
    timestamp = clean_text(created_at) or now_iso()
    progress: list[dict] = []
    for key, label, public_label in PROGRESS_STEPS:
        done = key == "submitted"
        progress.append(
            {
                "key": key,
                "label": label,
                "public_label": public_label,
                "done": done,
                "updated_at": timestamp if done else "",
                "note": "",
                "public_note": "",
            }
        )
    return progress


def normalize_progress(value: object, created_at: object = "") -> list[dict]:
    raw_items: list[dict] = []
    if isinstance(value, list):
        raw_items = [item for item in value if isinstance(item, dict)]
    elif value:
        try:
            parsed = json.loads(str(value))
            if isinstance(parsed, list):
                raw_items = [item for item in parsed if isinstance(item, dict)]
        except json.JSONDecodeError:
            raw_items = []

    by_key = {clean_text(item.get("key")): item for item in raw_items}
    normalized = default_progress(created_at)
    for item in normalized:
        existing = by_key.get(item["key"]) or {}
        item["done"] = bool(existing.get("done", item["done"]))
        item["updated_at"] = clean_text(existing.get("updated_at")) or item["updated_at"]
        item["note"] = clean_text(existing.get("note"))
        item["public_note"] = clean_text(existing.get("public_note"))
    return normalized


def public_progress(progress: list[dict]) -> list[dict]:
    return [
        {
            "key": clean_text(item.get("key")),
            "label": clean_text(item.get("public_label") or item.get("label")),
            "done": bool(item.get("done")),
            "updated_at": clean_text(item.get("updated_at")),
            "public_note": clean_text(item.get("public_note")),
        }
        for item in progress
    ]


def migrate_jsonl_to_database() -> None:
    submissions = read_jsonl("submissions.jsonl", limit=100000)
    matches = read_jsonl("matches.jsonl", limit=100000)
    intents = read_jsonl("intents.jsonl", limit=100000)
    if not submissions and not matches and not intents:
        return

    with db_connect() as conn:
        for item in reversed(submissions):
            db_execute(
                conn,
                """
                INSERT INTO submissions (submission_id, created_at, submission_json)
                VALUES (?, ?, ?)
                ON CONFLICT (submission_id) DO NOTHING
                """,
                (
                    clean_text(item.get("submission_id")) or uuid.uuid4().hex,
                    clean_text(item.get("created_at")) or now_iso(),
                    json_dumps(item.get("submission") or {}),
                ),
            )
        for item in reversed(matches):
            db_execute(
                conn,
                """
                INSERT INTO matches
                    (match_id, submission_id, created_at, ai_meta_json, results_json, submission_json)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT (match_id) DO NOTHING
                """,
                (
                    clean_text(item.get("match_id")) or uuid.uuid4().hex,
                    clean_text(item.get("submission_id")),
                    clean_text(item.get("created_at")) or now_iso(),
                    json_dumps(item.get("ai_meta") or {}),
                    json_dumps(item.get("results") or []),
                    json_dumps(item.get("submission") or {}),
                ),
            )
        for item in reversed(intents):
            db_execute(
                conn,
                """
                INSERT INTO intents
                    (intent_id, created_at, updated_at, status, agreement_version, submission_id,
                     contact_json, message, attachment_note, selected_result_json, followup_note,
                     query_code, public_note, progress_json, deal_amount, deal_note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (intent_id) DO NOTHING
                """,
                (
                    clean_text(item.get("intent_id")) or uuid.uuid4().hex,
                    clean_text(item.get("created_at")) or now_iso(),
                    clean_text(item.get("updated_at")) or clean_text(item.get("created_at")) or now_iso(),
                    clean_text(item.get("status")) or "待审核",
                    clean_text(item.get("agreement_version")),
                    clean_text(item.get("submission_id")),
                    json_dumps(item.get("contact") or {}),
                    clean_text(item.get("message")),
                    clean_text(item.get("attachment_note")),
                    json_dumps(item.get("selected_result") or {}),
                    clean_text(item.get("followup_note")),
                    clean_text(item.get("query_code")) or generate_unique_query_code(conn),
                    clean_text(item.get("public_note")) or f"已收到合作意向，{PUBLIC_RESPONSE_PROMISE}",
                    json_dumps(normalize_progress(item.get("progress"), item.get("created_at"))),
                    clean_text(item.get("deal_amount")),
                    clean_text(item.get("deal_note")),
                ),
            )


def db_count(table: str) -> int:
    if table not in {"submissions", "matches", "intents"}:
        return 0
    with db_connect() as conn:
        return int(first_value(db_execute(conn, f"SELECT COUNT(*) FROM {table}").fetchone()))


def save_submission(record: dict) -> None:
    with db_connect() as conn:
        db_execute(
            conn,
            """
            INSERT INTO submissions (submission_id, created_at, submission_json)
            VALUES (?, ?, ?)
            ON CONFLICT (submission_id) DO UPDATE SET
                created_at = excluded.created_at,
                submission_json = excluded.submission_json
            """,
            (
                clean_text(record.get("submission_id")),
                clean_text(record.get("created_at")) or now_iso(),
                json_dumps(record.get("submission") or {}),
            ),
        )


def save_match(record: dict) -> None:
    with db_connect() as conn:
        db_execute(
            conn,
            """
            INSERT INTO matches
                (match_id, submission_id, created_at, ai_meta_json, results_json, submission_json)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT (match_id) DO UPDATE SET
                submission_id = excluded.submission_id,
                created_at = excluded.created_at,
                ai_meta_json = excluded.ai_meta_json,
                results_json = excluded.results_json,
                submission_json = excluded.submission_json
            """,
            (
                clean_text(record.get("match_id")),
                clean_text(record.get("submission_id")),
                clean_text(record.get("created_at")) or now_iso(),
                json_dumps(record.get("ai_meta") or {}),
                json_dumps(record.get("results") or []),
                json_dumps(record.get("submission") or {}),
            ),
        )


def save_intent(intent: dict) -> None:
    timestamp = clean_text(intent.get("created_at")) or now_iso()
    with db_connect() as conn:
        query_code = clean_text(intent.get("query_code")) or generate_unique_query_code(conn)
        progress = normalize_progress(intent.get("progress") or intent.get("progress_json"), timestamp)
        db_execute(
            conn,
            """
            INSERT INTO intents
                (intent_id, created_at, updated_at, status, agreement_version, submission_id,
                 contact_json, message, attachment_note, selected_result_json, followup_note,
                 query_code, public_note, progress_json, deal_amount, deal_note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (intent_id) DO UPDATE SET
                created_at = excluded.created_at,
                updated_at = excluded.updated_at,
                status = excluded.status,
                agreement_version = excluded.agreement_version,
                submission_id = excluded.submission_id,
                contact_json = excluded.contact_json,
                message = excluded.message,
                attachment_note = excluded.attachment_note,
                selected_result_json = excluded.selected_result_json,
                followup_note = excluded.followup_note,
                query_code = excluded.query_code,
                public_note = excluded.public_note,
                progress_json = excluded.progress_json,
                deal_amount = excluded.deal_amount,
                deal_note = excluded.deal_note
            """,
            (
                clean_text(intent.get("intent_id")),
                timestamp,
                clean_text(intent.get("updated_at")) or timestamp,
                clean_text(intent.get("status")) or "待审核",
                clean_text(intent.get("agreement_version")),
                clean_text(intent.get("submission_id")),
                json_dumps(intent.get("contact") or {}),
                clean_text(intent.get("message")),
                clean_text(intent.get("attachment_note")),
                json_dumps(intent.get("selected_result") or {}),
                clean_text(intent.get("followup_note")),
                query_code,
                clean_text(intent.get("public_note")) or f"已收到合作意向，{PUBLIC_RESPONSE_PROMISE}",
                json_dumps(progress),
                clean_text(intent.get("deal_amount")),
                clean_text(intent.get("deal_note")),
            ),
        )


def decode_json_field(value: object, default: object) -> object:
    if not value:
        return default
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, (bytes, bytearray, memoryview)):
        value = bytes(value).decode("utf-8", errors="replace")
    try:
        return json.loads(str(value))
    except (json.JSONDecodeError, TypeError, ValueError):
        return default


def match_mode_label(ai_meta: dict) -> str:
    mode = clean_text(ai_meta.get("match_mode"))
    used_ai = bool(ai_meta.get("used_ai"))
    if mode == "quick":
        return "快速匹配"
    if used_ai:
        return "AI智能匹配"
    return "AI匹配-本地兜底"


def match_search_text(item: dict) -> str:
    submission = item.get("submission") or {}
    results = item.get("results") or []
    values: list[object] = [
        item.get("created_at"),
        item.get("match_mode_label"),
        item.get("ai_message"),
        submission.get("name"),
        submission.get("phone"),
        submission.get("company"),
        submission.get("title"),
        submission.get("tech_field"),
        submission.get("region"),
        submission.get("application_scene"),
        submission.get("summary"),
        submission.get("advantage"),
        submission.get("advantages"),
        submission.get("problem"),
        submission.get("maturity"),
        submission.get("ip_status"),
        submission.get("cooperation"),
        submission.get("extra_note"),
        submission.get("client_source"),
    ]
    for result in results:
        values.extend(
            [
                result.get("name"),
                result.get("tech_field"),
                result.get("demand_type"),
                result.get("region"),
                result.get("reason"),
                result.get("suggestion"),
            ],
        )
    return " ".join(clean_text(value) for value in values).lower()


def list_matches(limit: int = 120, keyword: str = "") -> list[dict]:
    keyword = clean_text(keyword).lower()
    with db_connect() as conn:
        rows = db_execute(
            conn,
            """
            SELECT match_id, submission_id, created_at, ai_meta_json, results_json, submission_json
            FROM matches
            ORDER BY created_at DESC
            """,
        ).fetchall()

    items: list[dict] = []
    for row in rows:
        item = dict(row)
        ai_meta = decode_json_field(item.pop("ai_meta_json", ""), {})
        results = decode_json_field(item.pop("results_json", ""), [])
        submission = decode_json_field(item.pop("submission_json", ""), {})
        if not isinstance(ai_meta, dict):
            ai_meta = {}
        if not isinstance(results, list):
            results = []
        if not isinstance(submission, dict):
            submission = {}
        item["ai_meta"] = ai_meta
        item["results"] = results[:5]
        item["submission"] = submission
        item["match_mode_label"] = match_mode_label(ai_meta)
        item["ai_message"] = clean_text(ai_meta.get("message"))
        if keyword and keyword not in match_search_text(item):
            continue
        items.append(item)
        if limit and len(items) >= limit:
            break
    return items


def get_match_record(match_id: str) -> dict:
    match_id = clean_text(match_id)
    with db_connect() as conn:
        row = db_execute(
            conn,
            """
            SELECT match_id, submission_id, created_at, ai_meta_json, results_json, submission_json
            FROM matches
            WHERE match_id = ?
            """,
            (match_id,),
        ).fetchone()
    if row is None:
        raise KeyError("未找到匹配记录")
    item = dict(row)
    item["ai_meta"] = decode_json_field(item.pop("ai_meta_json", ""), {})
    item["results"] = decode_json_field(item.pop("results_json", ""), [])
    item["submission"] = decode_json_field(item.pop("submission_json", ""), {})
    if not isinstance(item["ai_meta"], dict):
        item["ai_meta"] = {}
    if not isinstance(item["results"], list):
        item["results"] = []
    if not isinstance(item["submission"], dict):
        item["submission"] = {}
    item["match_mode_label"] = match_mode_label(item["ai_meta"])
    item["ai_message"] = clean_text(item["ai_meta"].get("message"))
    return item


def list_match_followups(match_id: str) -> dict[str, dict]:
    with db_connect() as conn:
        rows = db_execute(
            conn,
            """
            SELECT followup_id, match_id, submission_id, demand_id, status,
                   contact_note, project_progress, created_at, updated_at
            FROM match_followups
            WHERE match_id = ?
            """,
            (clean_text(match_id),),
        ).fetchall()
    return {clean_text(dict(row).get("demand_id")): dict(row) for row in rows}


def get_match_detail(store: DemandStore, match_id: str) -> dict:
    item = get_match_record(match_id)
    followups = list_match_followups(item["match_id"])
    enriched_results: list[dict] = []
    for result in item.get("results") or []:
        if not isinstance(result, dict):
            continue
        demand_id = clean_text(result.get("demand_id"))
        raw_demand = store.by_id(demand_id) or {}
        private = admin_demand_payload(raw_demand) if raw_demand else {}
        enriched_results.append(
            {
                **result,
                **private,
                "demand_id": demand_id,
                "followup": followups.get(demand_id)
                or {
                    "match_id": item["match_id"],
                    "submission_id": item.get("submission_id", ""),
                    "demand_id": demand_id,
                    "status": MATCH_FOLLOWUP_STATUSES[0],
                    "contact_note": "",
                    "project_progress": "",
                    "created_at": "",
                    "updated_at": "",
                },
            }
        )
    item["results"] = enriched_results
    return item


def save_match_followup(
    store: DemandStore,
    match_id: str,
    demand_id: str,
    status: str,
    contact_note: str = "",
    project_progress: str = "",
) -> dict:
    match_id = clean_text(match_id)
    demand_id = clean_text(demand_id)
    status = clean_text(status)
    if status not in MATCH_FOLLOWUP_STATUSES:
        raise ValueError("请选择有效的对接状态")
    match_record = get_match_record(match_id)
    candidate_ids = {
        clean_text(result.get("demand_id"))
        for result in (match_record.get("results") or [])
        if isinstance(result, dict)
    }
    if not demand_id or demand_id not in candidate_ids:
        raise ValueError("该需求不属于当前匹配记录")
    timestamp = now_iso()
    followup_id = hashlib.sha256(f"{match_id}:{demand_id}".encode("utf-8")).hexdigest()[:32]
    with db_connect() as conn:
        db_execute(
            conn,
            """
            INSERT INTO match_followups
                (followup_id, match_id, submission_id, demand_id, status,
                 contact_note, project_progress, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (match_id, demand_id) DO UPDATE SET
                status = excluded.status,
                contact_note = excluded.contact_note,
                project_progress = excluded.project_progress,
                updated_at = excluded.updated_at
            """,
            (
                followup_id,
                match_id,
                clean_text(match_record.get("submission_id")),
                demand_id,
                status,
                clean_text(contact_note),
                clean_text(project_progress),
                timestamp,
                timestamp,
            ),
        )
    return get_match_detail(store, match_id)


def list_intents(limit: int = 200, status: str = "", keyword: str = "") -> list[dict]:
    status = clean_text(status)
    keyword = clean_text(keyword).lower()
    clauses: list[str] = []
    params: list[str] = []
    if status and status != "全部":
        clauses.append("status = ?")
        params.append(status)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with db_connect() as conn:
        rows = db_execute(
            conn,
            f"""
            SELECT * FROM intents
            {where}
            ORDER BY created_at DESC
            """,
            params,
        ).fetchall()
    items: list[dict] = []
    for row in rows:
        item = dict(row)
        item["contact"] = decode_json_field(item.pop("contact_json", ""), {})
        item["selected_result"] = decode_json_field(item.pop("selected_result_json", ""), {})
        item["progress"] = normalize_progress(item.pop("progress_json", ""), item.get("created_at"))
        if keyword and keyword not in intent_search_text(item):
            continue
        items.append(item)
        if limit and len(items) >= limit:
            break
    return items


def row_to_intent(row: object) -> dict:
    item = dict(row)
    item["contact"] = decode_json_field(item.pop("contact_json", ""), {})
    item["selected_result"] = decode_json_field(item.pop("selected_result_json", ""), {})
    item["progress"] = normalize_progress(item.pop("progress_json", ""), item.get("created_at"))
    return item


def get_intent_detail(intent_id: str) -> dict:
    intent_id = clean_text(intent_id)
    with db_connect() as conn:
        row = db_execute(conn, "SELECT * FROM intents WHERE intent_id = ?", (intent_id,)).fetchone()
        if row is None:
            raise KeyError("未找到合作意向")
        logs = db_execute(
            conn,
            """
            SELECT old_status, new_status, note, operator, created_at
            FROM intent_status_logs
            WHERE intent_id = ?
            ORDER BY created_at DESC
            """,
            (intent_id,),
        ).fetchall()
    item = row_to_intent(row)
    item["status_logs"] = [dict(log) for log in logs]
    return item


def intent_search_text(item: dict) -> str:
    contact = item.get("contact") or {}
    selected = item.get("selected_result") or {}
    values = [
        item.get("created_at", ""),
        item.get("status", ""),
        item.get("query_code", ""),
        item.get("message", ""),
        item.get("attachment_note", ""),
        item.get("followup_note", ""),
        item.get("public_note", ""),
        item.get("deal_amount", ""),
        item.get("deal_note", ""),
        contact.get("name", ""),
        contact.get("phone", ""),
        contact.get("company", ""),
        contact.get("technology_summary", ""),
        selected.get("name", ""),
        selected.get("demand_id", ""),
        selected.get("score", ""),
        selected.get("reason", ""),
        selected.get("tech_field", ""),
        selected.get("demand_type", ""),
        selected.get("region", ""),
    ]
    return " ".join(clean_text(value) for value in values).lower()


def export_intents_xlsx(items: list[dict]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法导出 Excel。请先执行 pip install -r requirements.txt") from exc

    headers = [
        "创建时间",
        "当前状态",
        "查询码",
        "姓名",
        "手机号",
        "单位",
        "技术成果摘要",
        "匹配需求",
        "匹配分数",
        "技术领域",
        "需求类型",
        "所在地区",
        "意向投入",
        "AI匹配理由",
        "用户留言",
        "补充说明/相关链接",
        "跟进备注",
        "对外说明",
        "成交金额",
        "成交备注",
        "合作意向ID",
        "需求ID",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合作意向"
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in items:
        contact = item.get("contact") or {}
        selected = item.get("selected_result") or {}
        sheet.append(
            [
                item.get("created_at", ""),
                item.get("status", ""),
                item.get("query_code", ""),
                contact.get("name", ""),
                contact.get("phone", ""),
                contact.get("company", ""),
                contact.get("technology_summary", ""),
                selected.get("name", ""),
                selected.get("score", ""),
                selected.get("tech_field", ""),
                selected.get("demand_type", ""),
                selected.get("region", ""),
                selected.get("intended_price", ""),
                selected.get("reason", ""),
                item.get("message", ""),
                item.get("attachment_note", ""),
                item.get("followup_note", ""),
                item.get("public_note", ""),
                item.get("deal_amount", ""),
                item.get("deal_note", ""),
                item.get("intent_id", ""),
                selected.get("demand_id", ""),
            ]
        )

    widths = {
        "A": 20,
        "B": 16,
        "C": 14,
        "D": 18,
        "E": 22,
        "F": 34,
        "G": 36,
        "H": 12,
        "I": 26,
        "J": 18,
        "K": 24,
        "L": 14,
        "M": 48,
        "N": 28,
        "O": 30,
        "P": 30,
        "Q": 30,
        "R": 20,
        "S": 30,
        "T": 34,
        "U": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index in (3, 5, 20, 21):
            row[index - 1].number_format = "@"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def update_intent_status(
    intent_id: str,
    status: str,
    note: str,
    operator: str,
    *,
    public_note: str = "",
    progress: list[dict] | None = None,
    deal_amount: str | None = None,
    deal_note: str | None = None,
) -> dict:
    intent_id = clean_text(intent_id)
    status = clean_text(status)
    if status not in INTENT_STATUSES:
        raise ValueError("状态不在允许范围内")
    with db_connect() as conn:
        row = db_execute(conn, "SELECT * FROM intents WHERE intent_id = ?", (intent_id,)).fetchone()
        if row is None:
            raise KeyError("未找到合作意向")
        old_status = clean_text(row["status"])
        timestamp = now_iso()
        current = dict(row)
        progress_value = normalize_progress(progress if progress is not None else current.get("progress_json"), current.get("created_at"))
        public_note_value = clean_text(public_note) or clean_text(current.get("public_note"))
        deal_amount_value = clean_text(deal_amount) if deal_amount is not None else clean_text(current.get("deal_amount"))
        deal_note_value = clean_text(deal_note) if deal_note is not None else clean_text(current.get("deal_note"))
        db_execute(
            conn,
            """
            UPDATE intents
            SET status = ?, followup_note = ?, public_note = ?, progress_json = ?,
                deal_amount = ?, deal_note = ?, updated_at = ?
            WHERE intent_id = ?
            """,
            (
                status,
                clean_text(note),
                public_note_value,
                json_dumps(progress_value),
                deal_amount_value,
                deal_note_value,
                timestamp,
                intent_id,
            ),
        )
        db_execute(
            conn,
            """
            INSERT INTO intent_status_logs
                (log_id, intent_id, old_status, new_status, note, operator, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (uuid.uuid4().hex, intent_id, old_status, status, clean_text(note), clean_text(operator), timestamp),
        )
    updated = [item for item in list_intents(limit=1_000_000) if item.get("intent_id") == intent_id]
    return updated[0] if updated else {"intent_id": intent_id, "status": status}


def get_progress_by_query_code(query_code: str) -> dict:
    query_code = clean_text(query_code).upper()
    if not query_code:
        raise KeyError("请输入查询码")
    with db_connect() as conn:
        row = db_execute(conn, "SELECT * FROM intents WHERE UPPER(query_code) = ?", (query_code,)).fetchone()
        if row is None:
            raise KeyError("未找到该查询码对应的合作意向")
    item = row_to_intent(row)
    selected = item.get("selected_result") or {}
    return {
        "query_code": item.get("query_code", ""),
        "created_at": item.get("created_at", ""),
        "updated_at": item.get("updated_at", ""),
        "status": item.get("status", ""),
        "public_note": item.get("public_note", "") or f"已收到合作意向，{PUBLIC_RESPONSE_PROMISE}",
        "promise": PUBLIC_RESPONSE_PROMISE,
        "demand": {
            "name": selected.get("name", ""),
            "score": selected.get("score", ""),
            "tech_field": selected.get("tech_field", ""),
            "demand_type": selected.get("demand_type", ""),
            "region": selected.get("region", ""),
            "reason": clip(selected.get("reason", ""), 220),
            "scoring_source": selected.get("scoring_source", ""),
        },
        "progress": public_progress(item.get("progress") or []),
    }


def hash_password(password: str, salt: str | None = None, iterations: int = 220000) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
    return f"pbkdf2_sha256${iterations}${salt}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations_text, salt, expected = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        candidate = hash_password(password, salt=salt, iterations=int(iterations_text)).split("$", 3)[3]
        return hmac.compare_digest(candidate, expected)
    except Exception:
        return False


def ensure_admin_config() -> dict:
    env_username = clean_text(os.getenv("TECHNEXUS_ADMIN_USERNAME")) or "admin"
    env_password = clean_text(os.getenv("TECHNEXUS_ADMIN_PASSWORD"))
    if env_password:
        ensure_data_dir()
        config = {
            "username": env_username,
            "password_hash": hash_password(env_password),
            "created_at": now_iso(),
            "note": "由环境变量 TECHNEXUS_ADMIN_PASSWORD 生成管理员配置。",
        }
        ADMIN_CONFIG_FILE.write_text(json_dumps(config), encoding="utf-8")
        return config

    if not ADMIN_CONFIG_FILE.exists():
        ensure_data_dir()
        username = env_username
        password = secrets.token_urlsafe(12)
        config = {
            "username": username,
            "password_hash": hash_password(password),
            "created_at": now_iso(),
            "note": "首次自动生成管理员配置。请尽快运行 set_admin_password.py 修改密码。",
        }
        ADMIN_CONFIG_FILE.write_text(json_dumps(config), encoding="utf-8")
        if not os.getenv("TECHNEXUS_ADMIN_PASSWORD"):
            (DATA_DIR / "initial_admin_password.txt").write_text(
                f"TechNexus 初始管理员账号：{username}\nTechNexus 初始管理员密码：{password}\n请登录后尽快修改密码。\n",
                encoding="utf-8",
            )
        return config
    try:
        config = json.loads(ADMIN_CONFIG_FILE.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        config = {}
    if not isinstance(config, dict):
        config = {}
    if not config.get("username") or not config.get("password_hash"):
        ensure_data_dir()
        username = env_username
        password = secrets.token_urlsafe(12)
        config = {
            "username": username,
            "password_hash": hash_password(password),
            "created_at": now_iso(),
            "note": "管理员配置被修复。请尽快运行 set_admin_password.py 修改密码。",
        }
        ADMIN_CONFIG_FILE.write_text(json_dumps(config), encoding="utf-8")
        if not os.getenv("TECHNEXUS_ADMIN_PASSWORD"):
            (DATA_DIR / "initial_admin_password.txt").write_text(
                f"TechNexus 初始管理员账号：{username}\nTechNexus 初始管理员密码：{password}\n请登录后尽快修改密码。\n",
                encoding="utf-8",
            )
    return config


def login_matches_admin(username: str, password: str, config: dict) -> tuple[bool, str]:
    env_username = clean_text(os.getenv("TECHNEXUS_ADMIN_USERNAME")) or "admin"
    env_password = clean_text(os.getenv("TECHNEXUS_ADMIN_PASSWORD"))
    username = clean_text(username)
    password = clean_text(password)
    if env_password:
        return (
            username == env_username and hmac.compare_digest(password, env_password),
            env_username,
        )

    expected_user = clean_text(config.get("username"))
    expected_hash = clean_text(config.get("password_hash"))
    return (
        username == expected_user and verify_password(password, expected_hash),
        expected_user,
    )


def parse_cookie(header: str) -> dict[str, str]:
    cookies: dict[str, str] = {}
    for part in (header or "").split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        cookies[key.strip()] = value.strip()
    return cookies


class TechNexusHandler(BaseHTTPRequestHandler):
    store: DemandStore
    ai_config: dict
    admin_config: dict
    admin_sessions: dict[str, dict] = {}

    def log_message(self, format: str, *args: object) -> None:
        return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path in {"", "/"}:
            self.send_static(STATIC_DIR / "index.html")
            return
        if parsed.path.startswith("/static/"):
            relative = parsed.path.removeprefix("/static/")
            self.send_static(STATIC_DIR / relative)
            return
        if parsed.path == "/api/admin/session":
            username = self.current_admin()
            self.send_json({"authenticated": bool(username), "username": username})
            return
        if parsed.path == "/api/stats":
            self.send_json(
                {
                    **self.store.stats(),
                    **ai_status(self.ai_config),
                    "intent_count": db_count("intents"),
                    "match_count": db_count("matches"),
                }
            )
            return
        if parsed.path == "/api/public/demands":
            query = parse_qs(parsed.query)
            try:
                offset = max(0, int(query.get("offset", ["0"])[0] or 0))
                limit = max(1, min(60, int(query.get("limit", ["24"])[0] or 24)))
            except ValueError:
                self.send_error_json(HTTPStatus.BAD_REQUEST, "分页参数不正确")
                return
            items, total = self.store.search_page(keyword="", offset=offset, limit=limit)
            self.send_json(
                {
                    "items": [public_demand_payload(item) for item in items],
                    "total": total,
                    "offset": offset,
                    "limit": limit,
                }
            )
            return
        if parsed.path == "/api/intents":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            status = query.get("status", [""])[0]
            keyword = query.get("keyword", [""])[0]
            self.send_json({"items": list_intents(200, status=status, keyword=keyword), "statuses": INTENT_STATUSES})
            return
        if parsed.path == "/api/intents/detail":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            intent_id = query.get("intent_id", [""])[0]
            try:
                self.send_json({"intent": get_intent_detail(intent_id)})
            except KeyError as exc:
                self.send_error_json(HTTPStatus.NOT_FOUND, str(exc))
            return
        if parsed.path == "/api/intents/export":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            status = query.get("status", [""])[0]
            keyword = query.get("keyword", [""])[0]
            items = list_intents(10000, status=status, keyword=keyword)
            try:
                content = export_intents_xlsx(items)
            except RuntimeError as exc:
                self.send_error_json(HTTPStatus.INTERNAL_SERVER_ERROR, str(exc))
                return
            filename = f"TechNexus合作意向_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.send_binary(
                content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
                    "Cache-Control": "no-store",
                },
            )
            return
        if parsed.path == "/api/matches":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            keyword = query.get("keyword", [""])[0]
            self.send_json({"items": list_matches(120, keyword=keyword)})
            return
        if parsed.path == "/api/matches/detail":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            try:
                match = get_match_detail(self.store, query.get("match_id", [""])[0])
            except KeyError as exc:
                self.send_error_json(HTTPStatus.NOT_FOUND, str(exc))
                return
            self.send_json({"match": match, "statuses": MATCH_FOLLOWUP_STATUSES})
            return
        if parsed.path == "/api/demands":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            keyword = query.get("keyword", [""])[0]
            self.send_json({"items": self.store.search(keyword=keyword, limit=60)})
            return
        self.send_error_json(HTTPStatus.NOT_FOUND, "未找到页面")

    def do_POST(self) -> None:
        try:
            self.handle_post_request()
        except (BrokenPipeError, ConnectionResetError):
            return
        except Exception as exc:
            self.log_error("POST %s failed: %s", self.path, exc)
            try:
                self.send_error_json(
                    HTTPStatus.INTERNAL_SERVER_ERROR,
                    "服务器处理匹配请求时出现异常，请稍后重试；您填写的内容不会在前台公开。",
                )
            except (BrokenPipeError, ConnectionResetError, OSError):
                return

    def handle_post_request(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/admin/login":
            payload = self.read_json()
            username = clean_text(payload.get("username"))
            password = clean_text(payload.get("password"))
            ok, admin_username = login_matches_admin(username, password, self.admin_config)
            if not ok:
                self.send_error_json(HTTPStatus.UNAUTHORIZED, "账号或密码不正确")
                return
            token = secrets.token_urlsafe(32)
            self.admin_sessions[token] = {"username": admin_username, "expires_at": time.time() + SESSION_SECONDS}
            cookie = f"{SESSION_COOKIE}={token}; Path=/; Max-Age={SESSION_SECONDS}; HttpOnly; SameSite=Lax"
            self.send_json({"ok": True, "username": admin_username}, headers={"Set-Cookie": cookie})
            return

        if parsed.path == "/api/admin/logout":
            token = self.current_session_token()
            if token:
                self.admin_sessions.pop(token, None)
            cookie = f"{SESSION_COOKIE}=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax"
            self.send_json({"ok": True}, headers={"Set-Cookie": cookie})
            return

        if parsed.path == "/api/analyze-achievement":
            payload = self.read_json()
            submission = {
                clean_text(key): clean_text(value)
                for key, value in payload.items()
                if isinstance(value, (str, int, float, bool))
            }
            if not submission.get("title") and submission.get("achievement_name"):
                submission["title"] = submission["achievement_name"]
            if submission.get("achievement_text"):
                submission["achievement_text"] = clip(submission["achievement_text"], 30000)
            material = " ".join(
                clean_text(submission.get(field))
                for field in (
                    "achievement_text",
                    "title",
                    "summary",
                    "technical_route",
                    "problem",
                    "application_scene",
                    "indicators",
                    "evidence",
                )
            )
            if len(material) < 20:
                self.send_error_json(
                    HTTPStatus.BAD_REQUEST,
                    "请粘贴较完整的成果材料，或补充成果名称、技术原理和应用说明。",
                )
                return
            local_tags = extract_submission_tags_local(submission)
            tag_profile, tag_meta = extract_tags_with_ai(self.ai_config, submission, local_tags)
            capability_profile = normalize_technical_profile(
                tag_meta.get("capability_profile") or build_submission_technical_profile(submission)
            )
            self.send_json(
                {
                    "ok": True,
                    "used_ai": bool(tag_meta.get("used_ai")),
                    "source": clean_text(tag_meta.get("source") or "local"),
                    "message": clean_text(tag_meta.get("message")),
                    "structured_tags": compact_tag_payload(tag_profile),
                    "capability_profile": capability_profile,
                    "form_fields": capability_profile_form_fields(capability_profile),
                }
            )
            return

        if parsed.path == "/api/match":
            payload = self.read_json()
            provided_profile = payload.get("capability_profile") if isinstance(payload.get("capability_profile"), dict) else None
            provided_tags = payload.get("structured_tags") if isinstance(payload.get("structured_tags"), dict) else None
            provided_source = clean_text(payload.get("analysis_source") or "")
            match_mode = clean_text(payload.get("match_mode") or payload.get("_match_mode") or "ai").lower()
            if match_mode not in {"quick", "ai"}:
                match_mode = "ai"
            submission = {
                clean_text(k): clean_text(v)
                for k, v in payload.items()
                if clean_text(k) not in {
                    "match_mode",
                    "_match_mode",
                    "capability_profile",
                    "structured_tags",
                    "analysis_source",
                }
                and isinstance(v, (str, int, float, bool))
            }
            if not submission.get("title") and submission.get("achievement_name"):
                submission["title"] = submission["achievement_name"]
            if submission.get("achievement_text"):
                submission["achievement_text"] = clip(submission["achievement_text"], 30000)
            if clean_text(submission.get("client_source")) in {"网页端", "微信小程序"}:
                if not all(clean_text(submission.get(field)) for field in ("name", "phone", "company")):
                    self.send_error_json(HTTPStatus.BAD_REQUEST, "请填写姓名、手机号和单位，便于平台后续联系")
                    return
            submission_id = uuid.uuid4().hex
            record = {
                "submission_id": submission_id,
                "created_at": now_iso(),
                "submission": submission,
            }
            local_tag_profile = extract_submission_tags_local(submission)
            local_capability_profile = build_submission_technical_profile(submission)
            if provided_profile:
                tag_profile = merge_tag_profiles(local_tag_profile, provided_tags or {})
                capability_profile = merge_technical_profiles(local_capability_profile, provided_profile)
                tag_meta = {
                    "used_ai": "ai" in provided_source,
                    "source": provided_source or "provided",
                    "local_tags": compact_tag_payload(local_tag_profile),
                    "merged_tags": compact_tag_payload(tag_profile),
                    "capability_profile": capability_profile,
                    "message": "已使用提交前生成的成果能力画像召回候选需求。",
                }
            elif match_mode == "ai":
                tag_profile, tag_meta = extract_tags_with_ai(self.ai_config, submission, local_tag_profile)
                capability_profile = normalize_technical_profile(
                    tag_meta.get("capability_profile") or local_capability_profile
                )
            else:
                tag_profile = local_tag_profile
                capability_profile = local_capability_profile
                tag_meta = {
                    "used_ai": False,
                    "source": "local",
                    "local_tags": compact_tag_payload(local_tag_profile),
                    "merged_tags": compact_tag_payload(local_tag_profile),
                    "capability_profile": capability_profile,
                    "message": "已使用本地规则生成初步成果能力画像与召回标签。",
                }
            local_results = match_demands(
                self.store,
                submission,
                limit=20,
                candidate_limit=220,
                tags=tag_profile,
                capability_profile=capability_profile,
            )
            if match_mode == "quick":
                refined_results, ai_meta = use_quick_match(local_results)
            else:
                refined_results, ai_meta = refine_matches_with_ai(
                    self.ai_config,
                    submission,
                    local_results,
                    tag_profile=tag_profile,
                    tag_meta=tag_meta,
                    capability_profile=capability_profile,
                )
            ai_meta["tag_extraction"] = tag_meta
            ai_meta["structured_tags"] = compact_tag_payload(tag_profile)
            ai_meta["capability_profile"] = normalize_technical_profile(
                ai_meta.get("capability_profile") or capability_profile
            )
            ai_meta["local_candidate_count"] = len(local_results)
            results = [item for item in refined_results if int(item.get("score", 0) or 0) >= 45][:5]
            ai_meta["suppressed_low_relevance_count"] = max(0, len(refined_results) - len(results))
            save_submission(record)
            save_match(
                {
                    "match_id": uuid.uuid4().hex,
                    "submission_id": submission_id,
                    "created_at": now_iso(),
                    "ai_meta": ai_meta,
                    "submission": submission,
                    "results": results,
                }
            )
            self.send_json(
                {
                    "submission_id": submission_id,
                    "results": [public_demand_payload(item) for item in results],
                    "ai_meta": ai_meta,
                    "match_mode": match_mode,
                    **self.store.stats(),
                    **ai_status(self.ai_config),
                    "intent_count": db_count("intents"),
                    "match_count": db_count("matches"),
                }
            )
            return

        if parsed.path == "/api/intents":
            payload = self.read_json()
            if not payload.get("agreement"):
                self.send_error_json(HTTPStatus.BAD_REQUEST, "请先确认中介服务协议")
                return
            contact = payload.get("contact") or {}
            if not clean_text(contact.get("name")) or not clean_text(contact.get("phone")) or not clean_text(contact.get("company")):
                self.send_error_json(HTTPStatus.BAD_REQUEST, "请填写姓名、手机号和单位，便于后续人工撮合")
                return
            selected = payload.get("selected_result") or {}
            selected_demand = self.store.by_id(clean_text(selected.get("demand_id")))
            if selected_demand:
                selected = {**selected, **admin_demand_payload(selected_demand)}
            timestamp = now_iso()
            intent = {
                "intent_id": uuid.uuid4().hex,
                "created_at": timestamp,
                "status": "待审核",
                "agreement_version": AGREEMENT_VERSION,
                "submission_id": clean_text(payload.get("submission_id", "")),
                "contact": {clean_text(k): clean_text(v) for k, v in contact.items()},
                "message": clean_text(payload.get("message", "")),
                "attachment_note": clean_text(payload.get("extra_note") or payload.get("attachment_note", "")),
                "selected_result": selected,
                "updated_at": timestamp,
                "followup_note": "",
                "query_code": generate_unique_query_code(),
                "public_note": f"已收到合作意向，{PUBLIC_RESPONSE_PROMISE}",
                "progress": default_progress(timestamp),
                "deal_amount": "",
                "deal_note": "",
            }
            save_intent(intent)
            self.send_json(
                {
                    "ok": True,
                    "intent": {
                        **intent,
                        "selected_result": public_demand_payload(intent["selected_result"]),
                        "promise": PUBLIC_RESPONSE_PROMISE,
                    },
                }
            )
            return

        if parsed.path == "/api/progress/query":
            payload = self.read_json()
            try:
                progress = get_progress_by_query_code(clean_text(payload.get("query_code")))
            except KeyError as exc:
                self.send_error_json(HTTPStatus.NOT_FOUND, str(exc))
                return
            self.send_json({"ok": True, "progress": progress})
            return

        if parsed.path == "/api/intents/status":
            operator = self.require_admin()
            if not operator:
                return
            payload = self.read_json()
            try:
                intent = update_intent_status(
                    clean_text(payload.get("intent_id")),
                    clean_text(payload.get("status")),
                    clean_text(payload.get("note")),
                    operator,
                    public_note=clean_text(payload.get("public_note")),
                    progress=payload.get("progress") if isinstance(payload.get("progress"), list) else None,
                    deal_amount=clean_text(payload.get("deal_amount")) if "deal_amount" in payload else None,
                    deal_note=clean_text(payload.get("deal_note")) if "deal_note" in payload else None,
                )
            except KeyError as exc:
                self.send_error_json(HTTPStatus.NOT_FOUND, str(exc))
                return
            except ValueError as exc:
                self.send_error_json(HTTPStatus.BAD_REQUEST, str(exc))
                return
            self.send_json({"ok": True, "intent": intent, "items": list_intents(200)})
            return

        if parsed.path == "/api/matches/followup":
            if not self.require_admin():
                return
            payload = self.read_json()
            try:
                match = save_match_followup(
                    self.store,
                    clean_text(payload.get("match_id")),
                    clean_text(payload.get("demand_id")),
                    clean_text(payload.get("status")),
                    clean_text(payload.get("contact_note")),
                    clean_text(payload.get("project_progress")),
                )
            except KeyError as exc:
                self.send_error_json(HTTPStatus.NOT_FOUND, str(exc))
                return
            except ValueError as exc:
                self.send_error_json(HTTPStatus.BAD_REQUEST, str(exc))
                return
            self.send_json({"ok": True, "match": match, "statuses": MATCH_FOLLOWUP_STATUSES})
            return

        self.send_error_json(HTTPStatus.NOT_FOUND, "未找到接口")

    def current_session_token(self) -> str:
        cookies = parse_cookie(self.headers.get("Cookie", ""))
        return clean_text(cookies.get(SESSION_COOKIE))

    def current_admin(self) -> str:
        token = self.current_session_token()
        if not token:
            return ""
        session = self.admin_sessions.get(token)
        if not session:
            return ""
        if float(session.get("expires_at", 0)) < time.time():
            self.admin_sessions.pop(token, None)
            return ""
        return clean_text(session.get("username"))

    def require_admin(self) -> str:
        username = self.current_admin()
        if not username:
            self.send_error_json(HTTPStatus.UNAUTHORIZED, "请先登录后台")
            return ""
        return username

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        raw = self.rfile.read(length) if length else b"{}"
        try:
            payload = json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            payload = {}
        return payload if isinstance(payload, dict) else {}

    def send_json(
        self,
        payload: dict,
        status: HTTPStatus = HTTPStatus.OK,
        headers: Optional[dict[str, str]] = None,
    ) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status.value)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(body)

    def send_error_json(self, status: HTTPStatus, message: str) -> None:
        self.send_json({"ok": False, "message": message}, status=status)

    def send_binary(
        self,
        content: bytes,
        *,
        content_type: str,
        status: HTTPStatus = HTTPStatus.OK,
        headers: Optional[dict[str, str]] = None,
    ) -> None:
        self.send_response(status.value)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(content)

    def send_static(self, path: Path) -> None:
        try:
            resolved = path.resolve()
            if not str(resolved).startswith(str(STATIC_DIR.resolve())) or not resolved.exists() or not resolved.is_file():
                self.send_error_json(HTTPStatus.NOT_FOUND, "未找到静态文件")
                return
            content = resolved.read_bytes()
        except OSError:
            self.send_error_json(HTTPStatus.NOT_FOUND, "未找到静态文件")
            return
        content_type = mimetypes.guess_type(str(resolved))[0] or "application/octet-stream"
        if resolved.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif resolved.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif resolved.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        self.send_response(HTTPStatus.OK.value)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)


def run(host: str, port: int, open_browser: bool) -> None:
    init_database()
    admin_config = ensure_admin_config()
    store = DemandStore(DEMANDS_FILE)
    store.load()
    ai_config = load_ai_config()
    TechNexusHandler.store = store
    TechNexusHandler.ai_config = ai_config
    TechNexusHandler.admin_config = admin_config
    server = ThreadingHTTPServer((host, port), TechNexusHandler)
    url = f"http://{host}:{port}/"
    print(f"TechNexus 技术经理人试用版已启动：{url}")
    print(f"已读取需求库：{len(store.demands)} 条")
    print(f"匹配模式：{ai_status(ai_config)['ai_mode']}")
    print(f"后台管理员账号：{admin_config.get('username')}")
    print("按 Ctrl+C 可停止服务。")
    if open_browser:
        time.sleep(0.4)
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止服务。")
    finally:
        server.server_close()


def main() -> int:
    parser = argparse.ArgumentParser(description="TechNexus 技术经理人本地网页试用版")
    default_port = int(os.getenv("PORT") or os.getenv("TECHNEXUS_PORT") or 8010)
    parser.add_argument("--host", default=os.getenv("TECHNEXUS_HOST") or "127.0.0.1")
    parser.add_argument("--port", type=int, default=default_port)
    parser.add_argument("--no-browser", action="store_true", help="启动后不自动打开浏览器")
    args = parser.parse_args()
    run(args.host, args.port, open_browser=not args.no_browser)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
